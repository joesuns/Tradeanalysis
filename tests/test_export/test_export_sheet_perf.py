"""Export-E1: shared display transform + fast sheet write."""

import pandas as pd
from openpyxl import Workbook

from backend.export_wide import (
    _build_merged_display_df,
    _resolve_sheet_layout,
    _write_sheet_from_display,
)


def test_build_merged_display_enum_and_nulls():
    df = pd.DataFrame(
        {
            "ts_code": ["000001.SZ"],
            "stock_name": ["平安银行"],
            "macd_zone": ["bull"],
            "macd_divergence": [None],
            "pe_ttm": [None],
        }
    )
    daily_cols = ["ts_code", "stock_name", "macd_zone", "macd_divergence", "pe_ttm"]
    display, layout = _build_merged_display_df(df, daily_cols, weekly_cols=[])
    assert layout.n_basic == 3
    assert display.iloc[0]["macd_zone"] == "多头"
    assert display.iloc[0]["macd_divergence"] == "-"
    assert display.iloc[0]["pe_ttm"] == "N/A"


def test_signal_layout_is_subset_of_full():
    daily_cols = ["ts_code", "stock_name", "macd_zone", "dde_trend", "vol_signal"]
    weekly_cols = ["macd_zone"]
    merged = pd.DataFrame(
        {
            "ts_code": ["000001.SZ"],
            "stock_name": ["测试"],
            "macd_zone": ["bull"],
            "dde_trend": ["up"],
            "vol_signal": ["-"],
            "__w__macd_zone": ["bear"],
        }
    )
    full, _layout_full = _build_merged_display_df(merged, daily_cols, weekly_cols)
    layout_signal = _resolve_sheet_layout(
        ["ts_code", "stock_name", "macd_zone"],
        weekly_cols=["macd_zone"],
        available_columns=merged.columns,
    )
    signal = full[layout_signal.display_cols]
    assert len(signal.columns) < len(full.columns)
    assert signal.iloc[0]["macd_zone"] == "多头"
    assert signal.iloc[0]["__w__macd_zone"] == "空头"


def test_write_sheet_from_display_creates_data_rows():
    display = pd.DataFrame(
        {
            "ts_code": ["000001.SZ"],
            "stock_name": ["平安"],
            "macd_zone": ["多头"],
        }
    )
    layout = _resolve_sheet_layout(
        ["ts_code", "stock_name", "macd_zone"],
        weekly_cols=[],
        available_columns=list(display.columns),
    )
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet_from_display(wb, "测试", display, layout)
    ws = wb["测试"]
    assert ws.cell(row=3, column=1).value == "000001.SZ"
    assert ws.cell(row=3, column=2).value == "平安"
