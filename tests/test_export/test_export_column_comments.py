"""Tests for Excel export column comment glossary."""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from backend.export_column_comments import (
    DEFAULT_COMMENT_HEIGHT,
    DEFAULT_COMMENT_WIDTH,
    clear_comment_cache,
    comment_author,
    comment_box_size,
    format_column_comment,
    get_column_descriptions,
)
from backend.export_wide import _COL_NAMES, _attach_header_comment, _write_sheet_merged


_SKIP_KEYS = {"freq"}


def test_all_col_names_have_yaml_descriptions():
    descriptions = get_column_descriptions()
    missing = set(_COL_NAMES) - set(descriptions) - _SKIP_KEYS
    assert not missing, f"missing YAML descriptions: {sorted(missing)}"


def test_no_extra_yaml_keys_without_col_names():
    descriptions = get_column_descriptions()
    extra = set(descriptions) - set(_COL_NAMES)
    assert not extra, f"unknown YAML keys: {sorted(extra)}"


def test_format_column_comment_includes_footer():
    text = format_column_comment("dde_divergence")
    assert text is not None
    assert "TG" in text or "结构背离" in text
    assert "`-`" in text
    assert "N/A" in text


def test_format_column_comment_weekly_prefix():
    daily = format_column_comment("dde_divergence", weekly=False)
    weekly = format_column_comment("dde_divergence", weekly=True)
    assert daily is not None and weekly is not None
    assert weekly.startswith("【周线】")
    assert weekly.endswith(daily.split("\n\n")[-1])


def test_macd_turning_point_not_confused_with_zone():
    text = get_column_descriptions()["macd_turning_point"]
    assert "turning_point" in text or "金叉" in text
    assert "zone 逻辑" not in text


def test_comment_author():
    assert comment_author() == "Tradeanalysis"


def test_attach_header_comment_writes_openpyxl_comment():
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=2, column=5, value="DDE结构背离")
    _attach_header_comment(cell, "dde_divergence", weekly=False)
    assert isinstance(cell.comment, Comment)
    assert "结构背离" in cell.comment.text
    width, height = comment_box_size()
    assert cell.comment.width == width == DEFAULT_COMMENT_WIDTH
    assert cell.comment.height == height == DEFAULT_COMMENT_HEIGHT


def test_missing_yaml_degrades_without_comment(tmp_path):
    missing = tmp_path / "missing.yaml"
    clear_comment_cache()
    try:
        assert format_column_comment("dde_divergence", yaml_path=str(missing)) is None
        assert get_column_descriptions(yaml_path=str(missing)) == {}
    finally:
        clear_comment_cache()


def test_write_sheet_merged_persists_header_comments(tmp_path):
    df = pd.DataFrame(
        {
            "ts_code": ["000001.SZ"],
            "stock_name": ["平安银行"],
            "dde_divergence": ["-"],
        }
    )
    daily_cols = ["ts_code", "stock_name", "dde_divergence"]
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet_merged(wb, "测试", df, daily_cols, weekly_cols=[])
    out = tmp_path / "comments.xlsx"
    wb.save(out)

    loaded = load_workbook(out)
    ws = loaded["测试"]
    comment_cell = None
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        if cell.value == "DDE结构背离":
            comment_cell = cell
            break
    assert comment_cell is not None
    assert comment_cell.comment is not None
    assert "结构背离" in comment_cell.comment.text


def test_yaml_file_exists():
    path = Path(__file__).resolve().parents[2] / "docs" / "export" / "export-column-comments.yaml"
    assert path.is_file()
