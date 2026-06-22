"""Tests for portfolio stock sheet in Excel export."""
import io
import os
import tempfile

import duckdb
import pandas as pd
import pytest
from openpyxl import load_workbook

from backend.cli import _load_portfolio_stocks
from backend.export_wide import (
    _SIGNAL_ONLY,
    _resolve_portfolio_remarks,
    export_wide_to_excel,
)


# ---- _load_portfolio_stocks tests ----

def test_load_portfolio_valid_xlsx():
    """加载合法的持仓股列表 xlsx 文件"""
    # Create temp xlsx
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["stockcode", "stockname"])
    ws.append(["603986", "兆易创新"])
    ws.append(["002709", "天赐材料"])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        result = _load_portfolio_stocks(tmp.name)
        assert len(result) == 2
        assert result[0] == {"stockcode": "603986", "stockname": "兆易创新"}
        assert result[1] == {"stockcode": "002709", "stockname": "天赐材料"}
    finally:
        os.unlink(tmp.name)


def test_load_portfolio_case_insensitive_columns():
    """列名大小写不敏感匹配"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["StockCode", "StockName"])
    ws.append(["300346", "南大光电"])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        result = _load_portfolio_stocks(tmp.name)
        assert len(result) == 1
        assert result[0]["stockcode"] == "300346"
    finally:
        os.unlink(tmp.name)


def test_load_portfolio_strips_whitespace():
    """值自动 strip 前后空格"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["stockcode", "stockname"])
    ws.append([" 603986 ", " 兆易创新 "])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        result = _load_portfolio_stocks(tmp.name)
        assert result[0]["stockcode"] == "603986"
        assert result[0]["stockname"] == "兆易创新"
    finally:
        os.unlink(tmp.name)


def test_load_portfolio_missing_columns_warns_and_returns_empty(caplog):
    """缺少必需列时 WARNING + 返回空列表"""
    import openpyxl
    import logging
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["foo", "bar"])
    ws.append(["603986", "x"])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        with caplog.at_level(logging.WARNING):
            result = _load_portfolio_stocks(tmp.name)
        assert result == []
        assert "stockcode" in caplog.text.lower() or "column" in caplog.text.lower()
    finally:
        os.unlink(tmp.name)


def test_load_portfolio_file_not_found_returns_empty(caplog):
    """文件不存在时 WARNING + 返回空列表"""
    import logging
    with caplog.at_level(logging.WARNING):
        result = _load_portfolio_stocks("/nonexistent/path.xlsx")
    assert result == []


# ---- _resolve_portfolio_remarks tests ----

def test_resolve_remarks_normal_stock():
    """有当日数据的股票 → '正常'"""
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE dim_stock (stock_code VARCHAR, delist_date VARCHAR)")
    con.execute("INSERT INTO dim_stock VALUES ('603986', NULL)")
    con.execute("INSERT INTO dim_stock VALUES ('002709', NULL)")

    daily_stock_codes = {"603986", "002709"}
    portfolio_codes = ["603986", "002709"]
    trade_date = "20260620"

    remarks = _resolve_portfolio_remarks(con, portfolio_codes, trade_date, daily_stock_codes)
    assert remarks["603986"] == "正常"
    assert remarks["002709"] == "正常"
    con.close()


def test_resolve_remarks_delisted():
    """已退市股票 → '已退市'"""
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE dim_stock (stock_code VARCHAR, delist_date VARCHAR)")
    con.execute("INSERT INTO dim_stock VALUES ('000001', '20250101')")

    remarks = _resolve_portfolio_remarks(con, ["000001"], "20260620", set())
    assert remarks["000001"] == "已退市"
    con.close()


def test_resolve_remarks_not_in_db():
    """dim_stock 中不存在的股票 → '未入库'"""
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE dim_stock (stock_code VARCHAR, delist_date VARCHAR)")

    remarks = _resolve_portfolio_remarks(con, ["999999"], "20260620", set())
    assert remarks["999999"] == "未入库"
    con.close()


def test_resolve_remarks_suspended():
    """未退市但无当日行情 → '当日无数据（可能停牌）'"""
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE dim_stock (stock_code VARCHAR, delist_date VARCHAR)")
    con.execute("INSERT INTO dim_stock VALUES ('603986', NULL)")

    remarks = _resolve_portfolio_remarks(con, ["603986"], "20260620", set())
    assert "停牌" in remarks["603986"]
    con.close()


def test_resolve_remarks_mixed():
    """混合场景：正常 + 停牌 + 退市 + 未入库"""
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE dim_stock (stock_code VARCHAR, delist_date VARCHAR)")
    con.execute("INSERT INTO dim_stock VALUES ('A', NULL)")
    con.execute("INSERT INTO dim_stock VALUES ('B', NULL)")
    con.execute("INSERT INTO dim_stock VALUES ('C', '20250101')")
    # D not inserted → 未入库

    remarks = _resolve_portfolio_remarks(
        con, ["A", "B", "C", "D"], "20260620", {"A"},
    )
    assert remarks["A"] == "正常"
    assert "停牌" in remarks["B"]
    assert remarks["C"] == "已退市"
    assert remarks["D"] == "未入库"
    con.close()


# ---- Integration tests with export_wide_to_excel ----

def test_export_with_portfolio_creates_sheet(tmp_path):
    """有 portfolio_stocks 时生成持仓股分析 sheet 为第一 sheet"""
    db_path = str(tmp_path / "test.db")
    xlsx_path = str(tmp_path / "test.xlsx")

    con = duckdb.connect(db_path)

    # Setup minimal schema
    con.execute("CREATE TABLE dim_date (trade_date VARCHAR, is_week_end INTEGER)")
    con.execute("INSERT INTO dim_date VALUES ('20260620', 1)")

    con.execute("""
        CREATE TABLE dim_stock (
            stock_code VARCHAR, stock_name VARCHAR,
            exchange VARCHAR, industry VARCHAR,
            delist_date VARCHAR, is_st INTEGER
        )
    """)
    con.execute("INSERT INTO dim_stock VALUES ('603986', '兆易创新', 'SSE', '电子', NULL, 0)")
    con.execute("INSERT INTO dim_stock VALUES ('002709', '天赐材料', 'SZSE', '化工', NULL, 0)")
    con.execute("INSERT INTO dim_stock VALUES ('999999', '测试退市', 'SSE', '综合', '20250101', 0)")

    # Create views expected by export
    _create_minimal_analysis_views(con)

    con.close()

    portfolio = [
        {"stockcode": "603986", "stockname": "兆易创新"},
        {"stockcode": "002709", "stockname": "天赐材料"},
        {"stockcode": "999999", "stockname": "测试退市"},
        {"stockcode": "888888", "stockname": "未入库股"},
    ]

    result = export_wide_to_excel(
        db_path, "20260620", xlsx_path,
        filter_st=False, include_index=False,
        portfolio_stocks=portfolio,
    )

    assert result.row_count > 0

    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames
    assert sheet_names[0] == "持仓股分析"
    assert sheet_names[1] == "综合分析"
    assert sheet_names[2] == "个股分析"

    ws = wb["持仓股分析"]
    # Row 1: group header, Row 2: column names, Row 3+: data
    # Should have 4 data rows (2 normal + 2 missing)
    assert ws.max_row >= 5  # header rows + at least 3 data rows

    # Check remark column exists (last column — basic cols merged rows 1-2)
    last_col_header = ws.cell(row=1, column=ws.max_column).value
    assert last_col_header == "备注"

    wb.close()


def test_export_without_portfolio_no_sheet(tmp_path):
    """无 portfolio_stocks 时不生成持仓股分析 sheet（向后兼容）"""
    db_path = str(tmp_path / "test.db")
    xlsx_path = str(tmp_path / "test.xlsx")

    con = duckdb.connect(db_path)
    con.execute("CREATE TABLE dim_date (trade_date VARCHAR, is_week_end INTEGER)")
    con.execute("INSERT INTO dim_date VALUES ('20260620', 1)")
    _create_minimal_analysis_views(con)
    con.close()

    result = export_wide_to_excel(
        db_path, "20260620", xlsx_path,
        filter_st=False, include_index=False,
        portfolio_stocks=None,
    )

    wb = load_workbook(xlsx_path)
    assert "持仓股分析" not in wb.sheetnames
    assert wb.sheetnames[0] == "综合分析"
    wb.close()


def _create_minimal_analysis_views(con):
    """Create minimal v_ads_analysis_wide views for testing."""
    con.execute("""
        CREATE TABLE IF NOT EXISTS dwd_daily_quote (
            ts_code VARCHAR, trade_date VARCHAR, stock_code VARCHAR,
            stock_name VARCHAR, exchange VARCHAR, industry VARCHAR,
            is_st INTEGER, close DOUBLE, pct_chg DOUBLE, vol DOUBLE,
            amount DOUBLE, total_mv DOUBLE, pe_ttm DOUBLE, turnover_rate DOUBLE,
            volume_ratio DOUBLE, price_position_60d DOUBLE
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS dwd_weekly_quote (
            ts_code VARCHAR, trade_date VARCHAR, stock_code VARCHAR,
            stock_name VARCHAR, exchange VARCHAR, industry VARCHAR,
            is_st INTEGER, close DOUBLE, pct_chg DOUBLE, vol DOUBLE,
            amount DOUBLE, total_mv DOUBLE, pe_ttm DOUBLE, turnover_rate DOUBLE,
            volume_ratio DOUBLE
        )
    """)

    # Create empty plate tables to prevent load_plate_enrichment crash
    con.execute("""
        CREATE TABLE IF NOT EXISTS ods_plate_member (
            trade_date VARCHAR, source VARCHAR, con_code VARCHAR,
            board_ts_code VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS ods_plate_board (
            trade_date VARCHAR, source VARCHAR, board_ts_code VARCHAR,
            board_name VARCHAR
        )
    """)

    # Insert test data
    con.execute("""
        INSERT INTO dwd_daily_quote VALUES
        ('603986.SH', '20260620', '603986', '兆易创新', 'SSE', '电子', 0,
         100.0, 2.5, 1000, 5000, 500, 30, 3.5, 1.2, 75.5),
        ('002709.SZ', '20260620', '002709', '天赐材料', 'SZSE', '化工', 0,
         50.0, -1.0, 2000, 3000, 200, 25, 2.0, 0.8, 42.0)
    """)

    # Create views
    con.execute("CREATE OR REPLACE VIEW v_ads_analysis_wide_daily AS SELECT * FROM dwd_daily_quote")
    con.execute("CREATE OR REPLACE VIEW v_ads_analysis_wide_weekly AS SELECT * FROM dwd_weekly_quote")
