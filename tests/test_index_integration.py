"""End-to-end index pipeline integration tests."""
import pandas as pd
import pytest


@pytest.mark.integration
def test_index_export_column_order(db_with_schema, monkeypatch):
    """Verify index_name and index_category follow ts_code as columns 2 and 3,
    and pb immediately follows pe_ttm in the index overview sheet."""
    from openpyxl import Workbook
    import backend.export_index

    # Monkeypatch the view name to use a mock table we control
    monkeypatch.setattr(
        backend.export_index, "_INDEX_DAILY_VIEW", "mock_index_daily"
    )

    con = db_with_schema

    # Build mock data with all columns needed by _INDEX_BASIC_COLS and _INDEX_SIGNAL_COLS.
    # Use empty strings for nullable event-signal columns so DuckDB infers VARCHAR
    # rather than Int32 — otherwise apply_display_nulls' fillna("-") will fail.
    mock_data = {
        "freq": ["D"],
        "trade_date": ["20260601"],
        "ts_code": ["000001.SH"],
        "index_name": ["上证指数"],
        "index_category": ["core"],
        "close": [3000.0],
        "pct_chg": [1.5],
        "vol": [100000.0],
        "amount": [500000.0],
        "total_mv": [1000000000.0],
        "pe_ttm": [15.0],
        "pb": [1.5],
        "turnover_rate": [2.0],
        "volume_ratio": [1.0],
        "macd_divergence": ["[]"],
        "macd_zone": ["bull"],
        "macd_turning_point": [""],
        "macd_alert": [""],
        "macd_trend": ["up"],
        "macd_trend_strength": [0.05],
        "bias_ma5": [0.1],
        "bias_ma10": [0.05],
        "ma_alignment": ["多头强势"],
        "ma_turning_point": [""],
        "vol_zone": ["normal"],
        "vol_trend": ["flat"],
        "vol_divergence": ["[]"],
    }

    df = pd.DataFrame(mock_data)
    con.execute("CREATE TABLE mock_index_daily AS SELECT * FROM df")

    wb = Workbook()
    n = backend.export_index.export_index_sheet(con, "20260601", wb)

    assert n == 1, f"Should have 1 data row, got {n}"

    ws = wb["指数概览"]

    # Basic columns (ts_code, index_name, index_category) are merged across
    # rows 1-2 with their Chinese names on row 1.
    col1 = ws.cell(row=1, column=1).value
    col2 = ws.cell(row=1, column=2).value
    col3 = ws.cell(row=1, column=3).value

    assert col1 == "指数代码", f"Column 1 should be '指数代码', got '{col1}'"
    assert col2 == "指数名称", f"Column 2 should be '指数名称', got '{col2}'"
    assert col3 == "分类", f"Column 3 should be '分类', got '{col3}'"

    # Verify pb column appears immediately after pe_ttm column.
    # Both are basic columns with names on row 1.
    pb_col = None
    pe_col = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val == "市净率":
            pb_col = col_idx
        if val == "市盈率":
            pe_col = col_idx
    assert pb_col is not None, "pb should be visible as '市净率' in row 1"
    assert pe_col is not None, "pe_ttm should be visible as '市盈率' in row 1"
    assert pb_col == pe_col + 1, (
        f"pb should follow pe_ttm immediately, "
        f"got pe_ttm at col {pe_col}, pb at col {pb_col}"
    )


@pytest.mark.integration
@pytest.mark.slow
def test_index_full_pipeline(db_with_schema):
    """Fetch -> DIM -> DWD -> Calc -> Query — full pipeline for 000001.SH."""
    from backend.fetch.ods_index import (
        fetch_index_basic, fetch_index_daily,
    )
    from backend.etl.build_dim_index import build_dim_index
    from backend.etl.build_dwd_index import build_dwd_index_all
    from backend.etl.calc_index import calc_index_pipeline

    con = db_with_schema

    # 1. Fetch (requires tushare token, skipped in CI)
    import os
    token = os.environ.get("TUSHARE_TOKEN")
    if not token:
        pytest.skip("TUSHARE_TOKEN not set — skipping full pipeline test")

    from backend.fetch.client import get_client
    client = get_client()

    fetch_index_basic(client, con)
    fetch_index_daily(client, con)

    # 2. DIM
    n_dim = build_dim_index(con)
    assert n_dim >= 1

    # 3. DWD
    dwd_result = build_dwd_index_all(con)
    assert dwd_result["dwd_index_daily"] > 0

    # 4. Calc
    calc_date = con.execute(
        "SELECT MAX(trade_date) FROM dwd_index_daily"
    ).fetchone()[0]
    stats = calc_index_pipeline(con, calc_date)
    assert stats["index_macd_daily"]["calculated"] >= 1

    # 5. Query via view
    row = con.execute(f"""
        SELECT ts_code, close, dif, dea, macd_bar, macd_zone
        FROM v_ads_market_index_daily
        WHERE ts_code = '000001.SH' AND trade_date = ?
    """, [calc_date]).fetchone()
    assert row is not None
    assert row[1] is not None  # close
    assert row[4] is not None  # macd_bar
    assert row[5] in ("bull", "bear", None)  # macd_zone


@pytest.mark.integration
def test_index_export_sheet_populates(db_with_schema):
    """Index export sheet function doesn't crash with empty data."""
    from openpyxl import Workbook
    from backend.export_index import export_index_sheet

    wb = Workbook()
    n = export_index_sheet(db_with_schema, "20260601", wb)
    # Should return 0 for no data, not crash
    assert n >= 0
    # A sheet named "指数概览" should exist and have a placeholder
    ws_index = wb["指数概览"]
    assert ws_index.cell(row=1, column=1).value is not None
