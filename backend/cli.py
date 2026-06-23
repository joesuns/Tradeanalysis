"""CLI entry point for the Tradeanalysis data pipeline.

Usage:
    python -m backend.cli run
    python -m backend.cli run --date 20260604
    python -m backend.cli check
    python -m backend.cli fetch [--ts-code 000543.SZ] [--start 20150101]
    python -m backend.cli calc [--date 20260605] [--ts-code 000543.SZ]
    python -m backend.cli export --date 20260529 [--ts-code 000543.SZ]
    python -m backend.cli query --ts-code 000001.SZ
    python -m backend.cli prune [--keep 5]
    python -m backend.cli repair-weekly [--execute]
    python -m backend.cli backfill-dde-meta [--days 900] [--sync-dwd] [--recalc] [--workers 3]
    python -m backend.cli refresh-state [--date 20260609] [--dry-run]
    python -m backend.cli status
"""

import argparse
import logging
import os
import sys
import uuid
import warnings
from datetime import datetime

from backend.log_config import setup_logging, set_run_id

setup_logging()
logger = logging.getLogger(__name__)


from backend.cli_dates import ensure_trade_date
from backend.cli_dates import resolve_trade_date as _resolve_trade_date_value


def _resolve_trade_date(con, date: str = None) -> str:
    """解析分析日期：指定则用指定，不指定则用今天。"""
    return _resolve_trade_date_value(date)


def _ensure_trade_date(con, date: str) -> str:
    """确保 date 是交易日；不是则往前找最近交易日。"""
    return ensure_trade_date(con, date)


def _warn_export_coverage(db_path: str, trade_date: str, n_rows: int,
                        filter_st: bool, ts_codes=None):
    """Log WARNING when export rows are far below expected active stocks."""
    from backend.db.connection import get_connection

    con = get_connection(read_only=True)
    try:
        if ts_codes:
            expected = len(ts_codes)
        else:
            st_clause = " AND is_st = 0" if filter_st else ""
            expected = con.execute(f"""
                SELECT COUNT(*) FROM dim_stock
                WHERE list_date <= ?
                  AND (delist_date IS NULL OR delist_date >= ?)
                  {st_clause}
            """, [trade_date, trade_date]).fetchone()[0]
        ods_count = con.execute(
            "SELECT COUNT(*) FROM ods_daily WHERE trade_date = ?", [trade_date]
        ).fetchone()[0]
        threshold = int(expected * 0.8)
        if n_rows < threshold:
            logger.warning(
                "Export row count %d is far below expected ~%d for %s "
                "(ods_daily=%d on date). Check fetch/calc logs.",
                n_rows, expected, trade_date, ods_count,
            )
    finally:
        con.close()


# ── check ──

def cmd_check(_args):
    """Check environment connectivity: DuckDB + tushare."""
    from backend.db.connection import check_connectivity
    from backend.fetch.client import TushareClient

    # Orphan temp cleanup happens automatically in get_connection()
    db = check_connectivity()
    print(f"DuckDB: {db['duckdb']} (v{db['version']})")
    print(f"Disk free: {db['disk_free_mb']:,} MB | DB size: {db['db_size_mb']:,} MB")
    print(f"Temp disk free: {db.get('temp_disk_free_mb', '?'):,} MB")
    if db.get('backup_count', 0) > 0:
        print(f"Old backups: {db.get('backup_count', 0)} files "
              f"({db.get('backup_size_mb', 0):,} MB)")
    try:
        TushareClient().call("stock_basic", exchange="", list_status="L", limit=1)
        print("tushare: connected")
    except Exception as e:
        print(f"tushare: error — {e}")


# ── fetch ──

def cmd_fetch(args):
    """Pull ODS data into DuckDB.

    No --ts-code: date-batched parallel mode for full market.
    --ts-code: stock-batched mode, per-stock incremental detection.
    """
    from backend.db.connection import get_connection
    from backend.etl.error_handler import log_etl_end, log_etl_start
    from backend.fetch.client import TushareClient
    from backend.fetch.ods_daily import (
        fetch_by_date_range_parallel,
        fetch_stocks_incremental,
        get_all_active_codes,
    )

    client = TushareClient()
    con = get_connection()
    lid, t0 = log_etl_start(con, "cli_fetch")
    try:
        start = args.start or "20150101"
        end = args.end or "20991231"

        if args.ts_code:
            codes = args.ts_code if isinstance(args.ts_code, list) else [args.ts_code]
            logger.info("Stock-batched fetch: %d stocks, %s~%s", len(codes), start, end)
            n = fetch_stocks_incremental(client, con, codes, start=start, end=end)
            mode = "stock"
        else:
            codes = get_all_active_codes(con)
            logger.info("Date-batched fetch: %d active stocks, %s~%s", len(codes), start, end)
            n = fetch_by_date_range_parallel(
                start, end, workers=3, ts_codes=codes, con=con
            )
            mode = "date"

        # Plate (board/concept) data — low priority, skip on failure
        from backend.fetch.ods_plate import fetch_plate_data
        from datetime import datetime as _dt

        try:
            plate_date = _dt.now().strftime("%Y%m%d")
            plate_lid, plate_t0 = log_etl_start(con, "cli_fetch_plate")
            try:
                plate_results = fetch_plate_data(client, con, plate_date)
                total_members = sum(
                    r.get("n_members", 0) for r in plate_results.values()
                )
                log_etl_end(
                    con, plate_lid, "cli_fetch_plate", plate_t0, "success",
                    row_count=total_members,
                )
            except Exception as e:
                log_etl_end(
                    con, plate_lid, "cli_fetch_plate", plate_t0, "degraded",
                    error_msg=f"skipped: {e}",
                )
        except Exception:
            pass  # defensive: plate fetch must never block fetch

        # DC theme data — low priority, degrade on failure
        from backend.fetch.ods_plate import fetch_theme_data

        try:
            theme_lid, theme_t0 = log_etl_start(con, "cli_fetch_theme")
            try:
                theme_result = fetch_theme_data(client, con, plate_date)
                log_etl_end(
                    con, theme_lid, "cli_fetch_theme", theme_t0, "success",
                    row_count=theme_result.get("n_members", 0),
                )
            except Exception as e:
                log_etl_end(
                    con, theme_lid, "cli_fetch_theme", theme_t0, "degraded",
                    error_msg=f"skipped: {e}",
                )
        except Exception:
            pass  # defensive: theme fetch must never block fetch

        rows_written = int(n)
        completeness = {"mode": mode, "start": start, "end": end}
        if hasattr(n, "to_completeness"):
            completeness.update(n.to_completeness())
        log_etl_end(
            con, lid, "cli_fetch", t0, "success", row_count=rows_written,
            data_completeness=completeness,
        )
        logger.info("Fetch complete: %d ODS rows written", rows_written)
    except Exception:
        log_etl_end(con, lid, "cli_fetch", t0, "failed")
        raise
    finally:
        con.close()


# ── calc ──

def cmd_calc(args, skip_stale_fetch=False):
    """Compute DWS indicators.

    Auto-fetches missing warmup + stale latest-day ODS before calculating.
    No --ts-code: calculate all active stocks.
    With --refresh-spec: narrow FULL for indicators whose spec_version lags code.
    """
    from backend.db.connection import get_connection, run_checkpoint
    from backend.etl.orchestrator import run_calc

    from backend.etl.calc_preflight_context import pop_run_calc_handoff, pop_run_preflight_context

    con = get_connection()
    try:
        calc_date = None
        if getattr(args, "date", None):
            calc_date = _ensure_trade_date(
                con, _resolve_trade_date(con, args.date))
        elif getattr(args, "refresh_spec", None):
            calc_date = _ensure_trade_date(con, _resolve_trade_date(con, None))
        ts_codes = args.ts_code if args.ts_code else None

        refresh_spec = getattr(args, "refresh_spec", None)
        if refresh_spec:
            from backend.etl.calc_spec_refresh import cmd_refresh_spec

            cmd_refresh_spec(
                con, calc_date, refresh_spec, ts_codes,
                dry_run=getattr(args, "dry_run", False),
            )
            run_checkpoint(con)
            return

        preflight_ctx = pop_run_preflight_context()
        calc_handoff = pop_run_calc_handoff()
        indicator_filter = (
            calc_handoff.indicator_filter if calc_handoff else None
        )
        run_calc(
            con,
            ts_codes=ts_codes,
            auto_fetch=True,
            calc_date=calc_date,
            skip_stale_fetch=skip_stale_fetch,
            force=getattr(args, "force", False),
            preflight_ctx=preflight_ctx,
            indicator_filter=indicator_filter,
            calc_handoff=calc_handoff,
        )
    finally:
        con.close()


def _load_portfolio_stocks(filepath=None):
    """Load portfolio stock list from xlsx file.

    Auto-detects ``持仓股列表.xlsx`` in cwd if no path given.
    Returns list of dicts with keys ``stockcode`` and ``stockname``.
    Returns empty list on missing file, bad format, or missing columns.
    """
    if filepath is None:
        filepath = os.path.join(os.getcwd(), "持仓股列表.xlsx")

    if not os.path.exists(filepath):
        logger.warning("portfolio file not found: %s", filepath)
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not available; skipping portfolio load")
        return []

    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
    except Exception as exc:
        logger.warning("failed to open portfolio file %s: %s", filepath, exc)
        return []

    # Read header row
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        wb.close()
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    try:
        code_idx = header.index("stockcode")
        name_idx = header.index("stockname")
    except ValueError:
        logger.warning(
            "portfolio file missing required columns (stockcode, stockname); got %s",
            header,
        )
        wb.close()
        return []

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(code_idx, name_idx):
            continue
        code = str(row[code_idx]).strip() if row[code_idx] is not None else ""
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        if code and name:
            result.append({"stockcode": code, "stockname": name})

    wb.close()
    logger.info("loaded %d portfolio stocks from %s", len(result), filepath)
    return result


# ── export ──

def cmd_export(args):
    """Export analysis wide table to Excel (single day or --from/--to range)."""
    from backend.cli_dates import resolve_cli_dates
    from backend.db.connection import get_connection
    from backend.export_wide import default_export_path, export_wide_to_excel

    con = get_connection()
    try:
        dates = resolve_cli_dates(con, args, default_today=False)
    except ValueError as exc:
        print(f"Error: {exc}")
        raise SystemExit(1) from exc
    finally:
        con.close()

    if not dates:
        print("Error: export requires --date or --from/--to")
        raise SystemExit(1)

    db_path = args.db_path or "data/tradeanalysis.duckdb"
    ts_codes = args.ts_code if args.ts_code else None
    portfolio_stocks = _load_portfolio_stocks(getattr(args, "portfolio_file", None))
    outputs = []

    for trade_date in dates:
        out = args.output if len(dates) == 1 and args.output else default_export_path(trade_date)
        result = export_wide_to_excel(
            db_path,
            trade_date,
            out,
            filter_st=not args.include_st,
            include_index=not args.no_index,
            ts_codes=ts_codes,
            portfolio_stocks=portfolio_stocks or None,
        )
        _warn_export_coverage(
            db_path, trade_date, result.row_count,
            filter_st=not args.include_st, ts_codes=ts_codes,
        )
        outputs.append((trade_date, result.row_count, out))
        print(f"Exported {result.row_count} rows -> {out}")

    if len(outputs) > 1:
        print(f"export range complete: {len(outputs)} files")


# ── run ──

def _rebuild_dwd_for_run(con, codes: list[str], date: str, fetch_result) -> tuple:
    """Rebuild DWD after run fetch step.

    Uses FetchResult.changed_codes union find_stale_dwd_codes subset.
    Returns (dwd_result_dict, stale_codes_rebuilt, dwd_meta). Empty dict + [] + meta when skipped.
    """
    from backend.etl.orchestrator import find_stale_dwd_codes
    from backend.etl.pipeline_context import coerce_fetch_result

    from backend.etl.column_indicator_deps import (
        calc_affecting_changed_codes,
        fetch_blocks_dwd_calc,
    )

    fr = coerce_fetch_result(fetch_result)
    if fr.changed_field_events:
        changed = calc_affecting_changed_codes(fr.changed_field_events, date)
    elif fetch_blocks_dwd_calc(fr):
        changed = fr.changed_codes_for_date(date)
    else:
        changed = []
    stale = find_stale_dwd_codes(con, codes, date)
    stale_extra = sorted(set(stale) - set(changed))
    to_rebuild = sorted(set(changed) | set(stale))
    empty_meta = {"qfq_codes": [], "stale_extra_codes": stale_extra}
    if not to_rebuild:
        if fr.rows_written > 0:
            logger.info(
                "DWD already fresh for %s after fetch (written=%d ODS rows) — skip rebuild",
                date, fr.rows_written,
            )
        else:
            logger.info(
                "DWD fresh for %s — skip rebuild (%d stocks checked)",
                date, len(codes),
            )
        return {}, [], empty_meta

    from backend.etl.build_dwd import (
        find_stocks_needing_qfq_refresh,
        rebuild_dwd_for_stale,
    )

    qfq_codes = find_stocks_needing_qfq_refresh(con, to_rebuild, date)

    if fr.rows_written > 0 or changed:
        logger.info(
            "Rebuilding DWD for %d stocks on %s (changed=%d stale_dwd=%d written=%d)",
            len(to_rebuild), date, len(changed), len(stale), fr.rows_written,
        )
    else:
        logger.info(
            "DWD stale for %d/%d stocks on %s — rebuilding subset",
            len(to_rebuild), len(codes), date,
        )
    result = rebuild_dwd_for_stale(con, to_rebuild, date)
    meta = {"qfq_codes": qfq_codes, "stale_extra_codes": stale_extra}
    return result, to_rebuild, meta


def cmd_run(args):
    """One-command daily analysis: fetch → calc → export (single day or --from/--to range)."""
    from backend.cli_dates import resolve_cli_dates, run_date_range_loop
    from backend.db.connection import get_connection
    from backend.etl.error_handler import log_etl_end, log_etl_start

    con = get_connection()
    try:
        dates = resolve_cli_dates(con, args, default_today=True)
    except ValueError as exc:
        print(f"Error: {exc}")
        raise SystemExit(1) from exc
    finally:
        con.close()

    if len(dates) == 1:
        _cmd_run_single_day(args, dates[0])
        return

    logger.info("run date range: %d trading days (%s → %s)", len(dates), dates[0], dates[-1])
    con_range = get_connection()
    try:
        lid, t0 = log_etl_start(con_range, "cli_run_range")
        try:
            progress = run_date_range_loop(
                dates,
                lambda d: _cmd_run_single_day(args, d),
                continue_on_error=getattr(args, "continue_on_error", False),
                label="run",
            )
            log_etl_end(
                con_range, lid, "cli_run_range", t0, "success",
                row_count=len(progress["ok"]),
                data_completeness={"date_range_progress": progress},
            )
            print(
                f"run range complete: ok={len(progress['ok'])} failed={len(progress['failed'])}",
            )
        except Exception:
            log_etl_end(con_range, lid, "cli_run_range", t0, "failed")
            raise
    finally:
        con_range.close()


def _cmd_run_single_day(args, date: str):
    """Run fetch → DWD → calc → export for one analysis_date."""
    from backend.db.connection import get_connection
    from backend.etl.error_handler import log_etl_end, log_etl_start
    from backend.export_wide import (
        build_export_data_completeness,
        default_export_path,
        export_wide_to_excel,
    )
    from backend.fetch.client import TushareClient
    from backend.fetch.ods_daily import (
        fetch_by_date_range_parallel,
        fetch_stocks_incremental,
        get_all_active_codes,
    )

    db_path = args.db_path or "data/tradeanalysis.duckdb"
    ts_codes = args.ts_code if args.ts_code else None

    logger.info("=== Step 1/3: Fetching market data for %s ===", date)
    skip_dwd_calc = False
    fetch_result = None
    dwd_meta = {"qfq_codes": [], "stale_extra_codes": []}
    calc_handoff = None
    con = get_connection()
    try:
        codes = ts_codes or get_all_active_codes(con)
        lid, t0 = log_etl_start(con, "run_fetch")
        if ts_codes:
            client = TushareClient()
            fetch_result = fetch_stocks_incremental(
                client, con, codes, start=date, end=date, force_compare=True)
        else:
            fetch_result = fetch_by_date_range_parallel(
                date, date, workers=3, ts_codes=codes, con=con,
                skip_covered=False,
            )
        from backend.etl.pipeline_context import PipelineContext, coerce_fetch_result

        fetch_result = coerce_fetch_result(fetch_result)
        pipeline_ctx = PipelineContext.from_fetch(
            con, date, codes, fetch_result, mode="run",
            force_recalc=getattr(args, "force", False),
        )
        skip_dwd_calc = pipeline_ctx.skip_dwd_calc
        logger.info(
            "Fetch complete: written=%d unchanged=%d for %s",
            fetch_result.rows_written, fetch_result.rows_unchanged, date,
        )
        fetch_completeness = {
            "analysis_date": date,
            "stocks": len(codes),
            **fetch_result.to_completeness(),
        }
        log_etl_end(
            con, lid, "run_fetch", t0, "success",
            row_count=fetch_result.rows_written,
            data_completeness=fetch_completeness,
        )

        lid, t0 = log_etl_start(con, "run_rebuild_dwd")
        if pipeline_ctx.skip_dwd_calc:
            logger.info(
                "run pipeline shortcut: skip DWD+calc (unchanged ODS, prior calc exists)",
            )
            dwd_result, stale_rebuilt = {}, []
            log_etl_end(
                con, lid, "run_rebuild_dwd", t0, "success", row_count=0,
                data_completeness={
                    "analysis_date": date,
                    "skipped": True,
                    "pipeline_shortcut": True,
                    **pipeline_ctx.to_completeness(),
                },
            )
        else:
            dwd_result, stale_rebuilt, dwd_meta = _rebuild_dwd_for_run(
                con, codes, date, fetch_result,
            )
            from backend.etl.build_dwd import _dwd_rebuild_row_count

            rebuild_rows = _dwd_rebuild_row_count(dwd_result) if dwd_result else 0
            log_etl_end(
                con, lid, "run_rebuild_dwd", t0, "success", row_count=rebuild_rows,
                data_completeness={
                    "analysis_date": date,
                    "skipped": not dwd_result,
                    **fetch_result.to_completeness(),
                    "stale_count": len(stale_rebuilt),
                },
            )

        if dwd_result and stale_rebuilt and not pipeline_ctx.skip_dwd_calc:
            from backend.config import CALC_REUSE_REFRESH_CTX
            from backend.etl.calc_state_refresh import maybe_refresh_state_after_dwd_rebuild

            lid2, t02 = log_etl_start(con, "run_refresh_state")
            try:
                want_artifacts = CALC_REUSE_REFRESH_CTX
                refresh_result = maybe_refresh_state_after_dwd_rebuild(
                    con, stale_rebuilt, date, dwd_result,
                    return_artifacts=want_artifacts,
                )
                refresh_summary = None
                if refresh_result:
                    if want_artifacts and isinstance(refresh_result, tuple):
                        refresh_summary, tails_bundle = refresh_result
                        from backend.etl.calc_preflight_context import (
                            build_context_from_refresh,
                            set_run_preflight_context,
                        )
                        ctx = build_context_from_refresh(
                            calc_date=date,
                            stale_codes=stale_rebuilt,
                            summary=refresh_summary,
                            state_map=tails_bundle.get("state_map", {}),
                            tails_bundle=tails_bundle,
                        )
                        set_run_preflight_context(ctx)
                    else:
                        refresh_summary = refresh_result
                log_etl_end(
                    con, lid2, "run_refresh_state", t02, "success",
                    row_count=(refresh_summary or {}).get("records_written", 0),
                    data_completeness={
                        "analysis_date": date,
                        "stale_count": len(stale_rebuilt),
                        "preflight_ctx": bool(
                            want_artifacts and refresh_result
                            and isinstance(refresh_result, tuple)
                        ),
                        **(refresh_summary or {}),
                    },
                )
            except Exception:
                log_etl_end(con, lid2, "run_refresh_state", t02, "failed")
                raise

        if (
            dwd_result
            and stale_rebuilt
            and not pipeline_ctx.skip_dwd_calc
            and fetch_result is not None
        ):
            from backend.etl.backfill_dde_recalc import (
                maybe_invalidate_dde_after_column_patch,
            )

            maybe_invalidate_dde_after_column_patch(
                con, date, fetch_result, stale_rebuilt,
            )

        if not skip_dwd_calc and fetch_result is not None:
            from backend.etl.calc_preflight_context import RunCalcHandoff
            from backend.etl.column_indicator_deps import (
                active_route_keys,
                calc_routes_narrowed,
                resolve_run_calc_indicator_filter,
            )

            indicator_filter = resolve_run_calc_indicator_filter(
                con,
                fetch_result,
                changed_codes=fetch_result.changed_codes_for_date(date),
                stale_extra_codes=dwd_meta.get("stale_extra_codes", []),
                qfq_codes=dwd_meta.get("qfq_codes", []),
            )
            narrowed = calc_routes_narrowed(indicator_filter)
            if narrowed:
                logger.info(
                    "run calc route narrow: indicators=%s routes=%s",
                    indicator_filter,
                    active_route_keys(indicator_filter),
                )
            calc_handoff = RunCalcHandoff(
                indicator_filter=indicator_filter,
                calc_routes_narrowed=narrowed,
                active_routes=active_route_keys(indicator_filter),
            )

        # Plate (board/concept) fetch — low priority, degrade on failure
        from backend.fetch.ods_plate import fetch_plate_data

        try:
            plate_client = TushareClient()
            plate_lid, plate_t0 = log_etl_start(con, "run_fetch_plate")
            try:
                plate_results = fetch_plate_data(plate_client, con, date)
                total_members = sum(
                    r.get("n_members", 0) for r in plate_results.values()
                )
                log_etl_end(
                    con, plate_lid, "run_fetch_plate", plate_t0, "success",
                    row_count=total_members,
                )
            except Exception as e:
                log_etl_end(
                    con, plate_lid, "run_fetch_plate", plate_t0, "degraded",
                    error_msg=f"skipped: {e}",
                )
        except Exception:
            pass  # defensive: plate fetch must never block run

        # DC theme data — low priority, degrade on failure
        from backend.fetch.ods_plate import fetch_theme_data

        try:
            theme_lid, theme_t0 = log_etl_start(con, "run_fetch_theme")
            try:
                theme_result = fetch_theme_data(plate_client, con, date)
                log_etl_end(
                    con, theme_lid, "run_fetch_theme", theme_t0, "success",
                    row_count=theme_result.get("n_members", 0),
                )
            except Exception as e:
                log_etl_end(
                    con, theme_lid, "run_fetch_theme", theme_t0, "degraded",
                    error_msg=f"skipped: {e}",
                )
        except Exception:
            pass  # defensive: theme fetch must never block run

    finally:
        con.close()

    logger.info("=== Step 2/3: Computing indicators for %s ===", date)
    args.date = date
    if not skip_dwd_calc:
        if calc_handoff is not None:
            from backend.etl.calc_preflight_context import set_run_calc_handoff

            set_run_calc_handoff(calc_handoff)
        cmd_calc(args, skip_stale_fetch=True)
    else:
        logger.info("=== Step 2/3: Skipped calc (pipeline shortcut) ===")

    if not getattr(args, "skip_export", False):
        logger.info("=== Step 3/3: Exporting analysis for %s ===", date)
        args.output = default_export_path(date, args.output)

        portfolio_stocks = _load_portfolio_stocks(getattr(args, "portfolio_file", None))

        con = get_connection()
        try:
            lid, t0 = log_etl_start(con, "run_export")
            result = export_wide_to_excel(
                db_path,
                date,
                args.output,
                filter_st=not args.include_st,
                include_index=not args.no_index,
                ts_codes=ts_codes,
                portfolio_stocks=portfolio_stocks or None,
            )
            log_etl_end(
                con, lid, "run_export", t0, "success", row_count=result.row_count,
                data_completeness=build_export_data_completeness(
                    date, result.tradable_enrich,
                ),
            )
        finally:
            con.close()

        _warn_export_coverage(
            db_path, date, result.row_count, filter_st=not args.include_st, ts_codes=ts_codes,
        )
        print(f"Exported {result.row_count} rows -> {args.output}")
    else:
        logger.info("Skipping export (--skip-export)")
    logger.info("Done.")


# ── refresh ──

def cmd_refresh(args):
    """Force recalc scoped indicators (R1): fetch → DWD → FULL calc → optional export."""
    from backend.cli_dates import resolve_cli_dates, run_date_range_loop
    from backend.db.connection import get_connection
    from backend.etl.error_handler import log_etl_end, log_etl_start
    from backend.etl.refresh_pipeline import parse_indicator_filter, run_refresh_pipeline

    con = get_connection()
    try:
        dates = resolve_cli_dates(con, args, default_today=True)
    except ValueError as exc:
        print(f"Error: {exc}")
        raise SystemExit(1) from exc
    finally:
        con.close()

    ts_codes = args.ts_code if args.ts_code else None
    try:
        indicator_filter = parse_indicator_filter(getattr(args, "indicator", None))
    except ValueError as exc:
        print(f"Error: {exc}")
        raise SystemExit(1) from exc

    db_path = args.db_path or "data/tradeanalysis.duckdb"
    dry_run = getattr(args, "dry_run", False)

    if dry_run and len(dates) > 1:
        con = get_connection()
        try:
            from backend.etl.refresh_pipeline import (
                estimate_refresh_scope,
                resolve_refresh_routes,
            )
            from backend.fetch.ods_daily import get_all_active_codes

            codes = ts_codes or get_all_active_codes(con)
            routes = resolve_refresh_routes(indicator_filter)
            scope = estimate_refresh_scope(dates, codes, routes)
            print(
                f"refresh dry-run: dates={scope['dates']} stocks={scope['n_stocks']} "
                f"indicators={scope['indicators']} est_route_count={scope['est_route_count']}",
            )
        finally:
            con.close()
        return

    if len(dates) > 1 and not dry_run:
        from backend.etl.refresh_pipeline import (
            REFRESH_CONFIRM_ROUTE_THRESHOLD,
            estimate_refresh_scope,
            resolve_refresh_routes,
        )
        from backend.fetch.ods_daily import get_all_active_codes

        con = get_connection()
        try:
            codes = ts_codes or get_all_active_codes(con)
            routes = resolve_refresh_routes(indicator_filter)
            scope = estimate_refresh_scope(dates, codes, routes)
            if (
                scope["est_route_count"] > REFRESH_CONFIRM_ROUTE_THRESHOLD
                and not getattr(args, "confirm", False)
            ):
                print(
                    f"Error: Large refresh scope ({scope['est_route_count']} route-runs). "
                    f"Re-run with --confirm.",
                )
                raise SystemExit(1)
        finally:
            con.close()

    def _refresh_one(trade_date: str):
        con = get_connection()
        try:
            return run_refresh_pipeline(
                con, trade_date,
                ts_codes=ts_codes,
                indicator_filter=indicator_filter,
                do_export=getattr(args, "export", False),
                dry_run=dry_run,
                confirmed=getattr(args, "confirm", False),
                export_path=args.output,
                db_path=db_path,
            )
        finally:
            con.close()

    if len(dates) == 1:
        if dry_run:
            summary = _refresh_one(dates[0])
            _print_refresh_summary(dates[0], summary)
            return
        con = get_connection()
        try:
            lid, t0 = log_etl_start(con, "cli_refresh")
            try:
                summary = _refresh_one(dates[0])
                log_etl_end(
                    con, lid, "cli_refresh", t0, "success",
                    row_count=summary.get("calc", {}).get("calculated", 0),
                    data_completeness={"analysis_date": dates[0], **summary},
                )
            except Exception:
                log_etl_end(con, lid, "cli_refresh", t0, "failed")
                raise
        finally:
            con.close()
        _print_refresh_summary(dates[0], summary)
        return

    logger.info("refresh date range: %d days (%s → %s)", len(dates), dates[0], dates[-1])
    summaries = {}
    con = get_connection()
    lid, t0 = log_etl_start(con, "cli_refresh_range")
    try:
        def _one(d):
            summaries[d] = _refresh_one(d)

        progress = run_date_range_loop(
            dates, _one,
            continue_on_error=getattr(args, "continue_on_error", False),
            label="refresh",
        )
        log_etl_end(
            con, lid, "cli_refresh_range", t0, "success",
            row_count=len(progress["ok"]),
            data_completeness={
                "date_range_progress": {**progress, "summaries": {
                    d: s.get("calc", {}) for d, s in summaries.items()
                }},
            },
        )
        print(f"refresh range complete: ok={len(progress['ok'])} failed={len(progress['failed'])}")
        for d in progress["ok"]:
            calc = summaries[d].get("calc", {})
            print(f"  {d}: calculated={calc.get('calculated', 0)}")
    except Exception:
        log_etl_end(con, lid, "cli_refresh_range", t0, "failed")
        raise
    finally:
        con.close()


def _print_refresh_summary(date: str, summary: dict) -> None:
    if summary.get("dry_run"):
        scope = summary
        print(
            f"refresh dry-run: dates={scope['dates']} stocks={scope['n_stocks']} "
            f"indicators={scope['indicators']} est_route_count={scope['est_route_count']}",
        )
        return
    calc = summary.get("calc", {})
    print(
        f"refresh complete: date={date} calculated={calc.get('calculated', 0)} "
        f"routes={calc.get('routes', [])}",
    )
    if summary.get("export"):
        exp = summary["export"]
        print(f"Exported {exp['row_count']} rows -> {exp['path']}")


# ── backfill-state ──

def cmd_backfill_state(args):
    """Backfill dws_calc_state for stocks missing routing state (one-time FULL)."""
    from datetime import datetime

    from backend.db.connection import get_connection
    from backend.etl.calc_gate import assert_calc_date_ready, resolve_effective_calc_date
    from backend.etl.calc_state_backfill import backfill_calc_state, find_missing_state_keys
    from backend.config import CALC_STRICT_DATE
    from backend.db.schema import ensure_calc_state_table
    from backend.fetch.ods_daily import get_all_active_codes

    con = get_connection()
    try:
        ensure_calc_state_table(con)
        calc_date = args.date
        if calc_date:
            calc_date = _ensure_trade_date(con, _resolve_trade_date(con, calc_date))
        else:
            calc_date = datetime.now().strftime("%Y%m%d")
        if CALC_STRICT_DATE:
            assert_calc_date_ready(con, calc_date, strict=True)
        else:
            calc_date = resolve_effective_calc_date(con, calc_date, cap_to_ods=True)

        ts_codes = args.ts_code if args.ts_code else get_all_active_codes(con)
        gaps = find_missing_state_keys(con, ts_codes)
        n_keys = sum(len(v) for v in gaps.values())
        print(f"Missing state: {len(gaps)} stocks, {n_keys} indicator×freq keys")
        if not gaps:
            return
        summary = backfill_calc_state(con, list(gaps.keys()), calc_date)
        print(
            f"Backfill complete: {summary['stocks']} stocks, "
            f"{summary['indicators']} runs, {summary['calculated']} DWS rows"
        )
    finally:
        con.close()


# ── prune ──

def cmd_prune(args):
    """Prune superseded DWS snapshots, keeping the last N runs.

    Deletes only rows made obsolete by newer calc_date snapshots; the
    latest-per-key value for every (ts_code, trade_date) is always kept,
    so v_*_latest views are unchanged. Runs a CHECKPOINT afterwards to
    reclaim space within the database file.

    With --cleanup-backups: remove old pre-* DuckDB backup files
    instead of DWS pruning, retaining the most recent N (default 2,
    via PRUNE_KEEP_BACKUPS).
    """
    from backend.db.connection import get_connection, prune_dws_snapshots, run_checkpoint, cleanup_backup_files
    from backend.config import DUCKDB_PATH
    import os

    if args.cleanup_backups:
        data_dir = os.path.dirname(os.path.abspath(DUCKDB_PATH)) or "."
        result = cleanup_backup_files(data_dir, keep=args.keep_backups, dry_run=args.dry_run)
        if args.dry_run:
            if result["deleted"]:
                print(f"Would delete {len(result['deleted'])} backup(s) "
                      f"({result['freed_mb']:,} MB):")
                for name in result["deleted"]:
                    print(f"  - {name}")
                print(f"Would retain: {', '.join(result['retained']) or 'none'}")
            else:
                print(f"No backup files to clean (retained: {result.get('retained', [])})")
            return
        if result["deleted"]:
            print(f"Deleted {len(result['deleted'])} backup(s) "
                  f"({result['freed_mb']:,} MB freed)")
        else:
            print("No old backup files to delete")
        return

    con = get_connection()
    try:
        deleted = prune_dws_snapshots(con, keep_runs=args.keep)
        run_checkpoint(con)
        total = sum(deleted.values())
        for table, n in deleted.items():
            print(f"{table:30s} {n:>12,}")
        print(f"{'TOTAL':30s} {total:>12,} rows pruned (keep_runs={args.keep})")
    finally:
        con.close()


# ── refresh-state ──

def cmd_refresh_state(args):
    """Realign dws_calc_state fingerprints with current DWD tails (no DWS recalc).

    Use after one-off full-market DWD rebuild poisons append routing while DWS
    snapshots remain valid. Typical wall clock ~10-15 min for full market.

    Tail load uses isolated read-only connections (parallel when
    REFRESH_STATE_PARALLEL=1); write lock held only for upsert + checkpoint.
    """
    from datetime import datetime

    from backend.db.connection import get_connection, run_checkpoint
    from backend.etl.calc_gate import assert_calc_date_ready, resolve_effective_calc_date
    from backend.etl.calc_state_refresh import refresh_calc_state_fingerprints
    from backend.etl.error_handler import log_etl_end, log_etl_start
    from backend.config import CALC_STRICT_DATE
    from backend.db.schema import ensure_calc_state_table
    from backend.fetch.ods_daily import get_all_active_codes

    # Phase 1: resolve date/codes without holding a write lock through tail SQL.
    setup_con = get_connection(read_only=True)
    try:
        calc_date = args.date
        if calc_date:
            calc_date = _ensure_trade_date(setup_con, _resolve_trade_date(setup_con, calc_date))
        else:
            calc_date = datetime.now().strftime("%Y%m%d")
        if CALC_STRICT_DATE:
            assert_calc_date_ready(setup_con, calc_date, strict=True)
        else:
            calc_date = resolve_effective_calc_date(setup_con, calc_date, cap_to_ods=True)
        ts_codes = args.ts_code if args.ts_code else get_all_active_codes(setup_con)
    finally:
        setup_con.close()

    logger.info(
        "refresh-state: %d stocks, calc_date=%s, dry_run=%s, tail_load=isolated",
        len(ts_codes), calc_date, args.dry_run,
    )

    # Phase 2: etl audit row (short write connection).
    audit_con = get_connection()
    try:
        ensure_calc_state_table(audit_con)
        lid, t0 = log_etl_start(audit_con, "cli_refresh_state")
    finally:
        audit_con.close()

    # Phase 3: isolated parallel read load + fingerprint scan (+ upsert if needed).
    try:
        summary = refresh_calc_state_fingerprints(
            None,
            ts_codes,
            calc_date,
            dry_run=args.dry_run,
            isolated_tail_load=True,
        )
    except Exception:
        audit_con = get_connection()
        try:
            log_etl_end(audit_con, lid, "cli_refresh_state", t0, "failed")
        finally:
            audit_con.close()
        raise

    # Phase 4: checkpoint + close audit (short write connection).
    audit_con = get_connection()
    try:
        if not args.dry_run:
            run_checkpoint(audit_con)
        log_etl_end(
            audit_con, lid, "cli_refresh_state", t0, "success",
            row_count=summary.get("records_written", 0),
            data_completeness=summary,
        )
    finally:
        audit_con.close()

    print(
        f"refresh-state {'(dry-run) ' if args.dry_run else ''}complete: "
        f"{summary['stocks']} stocks, "
        f"updated={summary['keys_updated']}, "
        f"unchanged={summary['keys_unchanged']}, "
        f"skipped={summary['keys_skipped']}, "
        f"written={summary['records_written']}, "
        f"elapsed={summary['elapsed_sec']}s, "
        f"tail_load={summary.get('tail_load_mode', 'isolated')}"
    )
    if not args.dry_run:
        print(
            f"Preflight after refresh: skip={summary['preflight_skip']}, "
            f"full={summary['preflight_full']}, "
            f"append={summary['preflight_append']}, "
            f"chunk_stocks={summary['chunk_stocks']}"
        )


# ── backfill-dde-meta ──

def cmd_backfill_dde_meta(args):
    """Backfill net_amount_dc + circ_mv for B4 weekly DDE trend (ops)."""
    from backend.db.connection import get_connection, run_checkpoint
    from backend.fetch.backfill_dde_meta import (
        MONEYFLOW_DC_MIN,
        backfill_dde_meta_ods,
        resolve_backfill_range,
    )
    from backend.etl.sync_dwd_dde_meta import sync_dwd_dde_meta
    from backend.fetch.client import TushareClient

    end = args.end or args.date or datetime.now().strftime("%Y%m%d")
    since = args.since or MONEYFLOW_DC_MIN
    start, end = resolve_backfill_range(end, days=args.days, since=since)
    workers = args.workers
    sync_batch = args.sync_dwd_batch if args.sync_dwd else 0
    want_recalc = args.recalc or args.recalc_only
    recalc_stats = None

    con = get_connection()
    try:
        if args.recalc_only:
            if args.dry_run:
                from backend.etl.backfill_dde_recalc import prepare_dde_weekly_recalc
                recalc_stats = prepare_dde_weekly_recalc(
                    con, end, ts_codes=args.ts_code, dry_run=True,
                )
                print(f"DDE recalc (dry-run): {recalc_stats}")
                return
            from backend.etl.backfill_dde_recalc import prepare_dde_weekly_recalc
            recalc_stats = prepare_dde_weekly_recalc(
                con, end, ts_codes=args.ts_code, dry_run=False,
            )
            run_checkpoint(con)
        elif not args.sync_dwd_only:
            client = TushareClient()

            def _sync_hook(c):
                sync_dwd_dde_meta(c, ts_codes=args.ts_code, since=since)

            stats = backfill_dde_meta_ods(
                con, client, args.ts_code, start, end,
                dry_run=args.dry_run,
                workers=workers,
                sync_dwd_batch=sync_batch,
                on_batch_sync=_sync_hook if sync_batch else None,
            )
            print(f"ODS backfill: {stats}")

            if args.dry_run:
                if want_recalc:
                    from backend.etl.backfill_dde_recalc import prepare_dde_weekly_recalc
                    recalc_stats = prepare_dde_weekly_recalc(
                        con, end, ts_codes=args.ts_code, dry_run=True,
                    )
                    print(f"DDE recalc (dry-run): {recalc_stats}")
                return

            if args.sync_dwd:
                if not (sync_batch > 0):
                    sync_stats = sync_dwd_dde_meta(
                        con, ts_codes=args.ts_code, since=since,
                    )
                    print(f"DWD sync: {sync_stats}")
        else:
            if args.dry_run:
                return
            sync_stats = sync_dwd_dde_meta(con, ts_codes=args.ts_code, since=since)
            print(f"DWD sync: {sync_stats}")

        if want_recalc and not args.dry_run and not args.recalc_only:
            from backend.etl.backfill_dde_recalc import prepare_dde_weekly_recalc
            recalc_stats = prepare_dde_weekly_recalc(
                con, end, ts_codes=args.ts_code, dry_run=False,
            )
            run_checkpoint(con)
    finally:
        con.close()

    if recalc_stats is not None and not args.dry_run:
        from backend.etl.backfill_dde_recalc import run_calc_force_hard_subprocess
        print(f"DDE recalc prepare: {recalc_stats}")
        run_calc_force_hard_subprocess(end, ts_codes=args.ts_code)
        print("DDE recalc calc: subprocess ok")


# ── repair-dde-trend ──

def cmd_repair_dde_trend(args):
    """Invalidate DDE trend routing+DWS and CALC_FORCE_HARD recalc (ops)."""
    from backend.db.connection import get_connection, run_checkpoint
    from backend.etl.backfill_dde_recalc import (
        prepare_dde_daily_recalc,
        prepare_dde_weekly_recalc,
        run_calc_force_hard_subprocess,
    )
    from backend.etl.error_handler import log_etl_end, log_etl_start

    calc_date = args.date
    con = get_connection()
    lid, t0 = log_etl_start(con, "repair_dde_trend")
    stats = {"freq": args.freq}
    try:
        if args.freq in ("daily", "both"):
            stats["daily"] = prepare_dde_daily_recalc(
                con, calc_date, ts_codes=args.ts_code, dry_run=args.dry_run,
                purge_history=getattr(args, "purge_history", False),
            )
        if args.freq in ("weekly", "both"):
            stats["weekly"] = prepare_dde_weekly_recalc(
                con, calc_date, ts_codes=args.ts_code, dry_run=args.dry_run,
            )
        if args.dry_run:
            log_etl_end(
                con, lid, "repair_dde_trend", t0, "success",
                data_completeness={**stats, "dry_run": True},
            )
            print(stats)
            return
        run_checkpoint(con)
    except Exception:
        log_etl_end(con, lid, "repair_dde_trend", t0, "failed")
        raise
    finally:
        con.close()

    if args.freq in ("daily", "both"):
        effective_date = stats["daily"]["calc_date"]
    else:
        effective_date = stats["weekly"]["calc_date"]

    run_calc_force_hard_subprocess(effective_date, ts_codes=args.ts_code)

    con = get_connection()
    try:
        log_etl_end(
            con, lid, "repair_dde_trend", t0, "success",
            data_completeness=stats,
        )
    finally:
        con.close()
    print("repair-dde-trend: calc subprocess ok")


# ── repair-weekly ──

def cmd_repair_weekly(args):
    """Repair weekly data after the date_trunc('week') partition fix.

    Default is a read-only dry-run that previews wrongly-marked week-ends and
    orphan DWS rows. Pass --execute to rebuild dim_date + dwd_weekly_quote and
    delete orphan rows. After executing, run `calc` to refresh stale week-end
    values (fingerprint auto-skips unchanged weeks).
    """
    from backend.db.connection import get_connection
    from backend.etl.repair_weekly import repair_weekly

    con = get_connection()
    try:
        res = repair_weekly(con, dry_run=not args.execute)
        print(f"Wrongly-marked week-ends: {len(res['wrongly_marked'])}")
        print(f"Newly-correct week-ends:  {len(res['newly_marked'])}")
        print("Orphan rows per weekly DWS table:")
        for tbl, n in res["orphans"].items():
            print(f"  {tbl:30s} {n:>12,}")
        if res["executed"]:
            print("EXECUTED — deleted orphan rows:")
            for tbl, n in res["deleted"].items():
                print(f"  {tbl:30s} {n:>12,}")
            print(f"Weekly calc_state invalidated: {res.get('weekly_state_invalidated', 0):,}")
            print("NOTE: run `python -m backend.cli calc` to refresh stale week-end values.")
        else:
            print("DRY-RUN (no changes). Re-run with --execute to apply.")
    finally:
        con.close()


# ── query / status ──

def cmd_query(args):
    """Query DWS indicators for a stock."""
    from backend.db.connection import get_connection
    con = get_connection(read_only=True)
    try:
        view = f"v_dws_macd_{args.freq}_latest"
        sql = (
            f"SELECT * FROM {view} "
            f"WHERE ts_code = ? "
            f"AND trade_date = (SELECT MAX(trade_date) FROM {view} WHERE ts_code = ?)"
        )
        row = con.execute(sql, (args.ts_code, args.ts_code)).fetchone()
        if row:
            cols = [d[0] for d in con.description]
            for c, v in zip(cols, row):
                print(f"{c}: {v}")
        else:
            print(f"No data for {args.ts_code}")
    finally:
        con.close()


def cmd_status(_args):
    """Show database table statistics."""
    from backend.db.connection import get_connection

    con = get_connection(read_only=True)
    try:
        tables = [
            "ods_daily", "ods_daily_basic", "ods_moneyflow",
            "dwd_daily_quote", "dwd_weekly_quote",
            "dws_macd_daily", "dws_ma_daily", "dws_kpattern_daily",
            "dws_dde_daily", "dws_volume_daily", "dws_price_position_daily",
        ]
        for table in tables:
            try:
                cnt = con.execute(
                    f"SELECT COUNT(*) FROM {table}"
                ).fetchone()[0]
                latest = con.execute(
                    f"SELECT MAX(trade_date) FROM {table}"
                ).fetchone()[0]
                print(f"{table:30s} {cnt:>12,}  {latest or 'N/A'}")
            except Exception:
                print(f"{table:30s}  (not found)")
    finally:
        con.close()


# ── main ──

DEPRECATED_OPS_COMMANDS = frozenset({
    "backfill-state",
    "backfill-dde-meta",
    "repair-dde-trend",
    "refresh-state",
    "prune",
    "repair-weekly",
})


def _warn_deprecated_top_level(command: str) -> None:
    warnings.warn(
        f"Top-level '{command}' is deprecated; use "
        f"'python -m backend.cli ops {command}'",
        DeprecationWarning,
        stacklevel=2,
    )


def _add_backfill_state_args(p):
    p.add_argument("--date", help="Calc date YYYYMMDD (default: today)")
    p.add_argument("--ts-code", nargs="+", help="Stock codes (default: all active)")


def _add_refresh_state_args(p):
    p.add_argument("--date", help="updated_calc_date tag YYYYMMDD (default: today)")
    p.add_argument("--ts-code", nargs="+", help="Stock codes (default: all active)")
    p.add_argument("--dry-run", action="store_true",
                   help="Compute diff only; do not write dws_calc_state")


def _add_prune_args(p):
    from backend.config import DWS_PRUNE_KEEP_RUNS
    p.add_argument("--keep", type=int, default=DWS_PRUNE_KEEP_RUNS,
                   help=f"Number of most recent calc runs to retain "
                        f"(default {DWS_PRUNE_KEEP_RUNS}, env DWS_PRUNE_KEEP_RUNS)")
    p.add_argument("--cleanup-backups", action="store_true",
                   help="Remove old pre-* DuckDB backup files instead of DWS pruning")
    p.add_argument("--keep-backups", type=int, default=None,
                   help="Backup files to retain (default: PRUNE_KEEP_BACKUPS=2)")
    p.add_argument("--dry-run", action="store_true",
                   help="Preview cleanup-backups without deleting")


def _add_repair_weekly_args(p):
    p.add_argument("--execute", action="store_true",
                   help="Apply changes (default: dry-run preview only)")


def _add_repair_dde_trend_args(p):
    p.add_argument("--date", required=True, help="Calc date YYYYMMDD")
    p.add_argument("--freq", choices=["daily", "weekly", "both"], default="daily")
    p.add_argument("--ts-code", nargs="+")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--purge-history",
        action="store_true",
        help="With --ts-code: delete ALL dde/daily DWS rows for those stocks before recalc",
    )


def _add_backfill_dde_meta_args(p):
    p.add_argument("--date", help="End date YYYYMMDD (default: today)")
    p.add_argument("--end", help="Alias for --date")
    p.add_argument("--since", default="20230911", help="moneyflow_dc min date")
    p.add_argument("--days", type=int, default=900, help="Calendar-day lookback")
    p.add_argument("--ts-code", nargs="+", help="Stock subset (default: all active, excl BSE)")
    p.add_argument("--dry-run", action="store_true",
                     help="Count gaps only, no API/DWD writes")
    p.add_argument("--sync-dwd", action="store_true",
                     help="After ODS backfill, sync ODS→DWD")
    p.add_argument("--sync-dwd-only", action="store_true",
                     help="Skip ODS API; DWD sync only")
    p.add_argument("--workers", type=int, default=3,
                     help="Parallel day-chunk workers (default 3)")
    p.add_argument("--sync-dwd-batch", type=int, default=50,
                     help="DWD sync every N trading days (requires --sync-dwd)")
    p.add_argument("--recalc", action="store_true",
                     help="After ODS+DWD: refresh-state → invalidate dde weekly → calc")
    p.add_argument("--recalc-only", action="store_true",
                     help="Skip ODS API; DDE weekly recalc closure only")


def _register_ops_subparsers(ops_sp):
    p = ops_sp.add_parser(
        "backfill-state",
        help="Backfill dws_calc_state for stocks missing routing state",
    )
    _add_backfill_state_args(p)

    p = ops_sp.add_parser(
        "backfill-dde-meta",
        help="Backfill net_amount_dc + circ_mv for B4 weekly DDE trend",
    )
    _add_backfill_dde_meta_args(p)

    p = ops_sp.add_parser(
        "repair-dde-trend",
        help="Invalidate DDE trend routing+DWS and CALC_FORCE_HARD recalc",
    )
    _add_repair_dde_trend_args(p)

    p = ops_sp.add_parser(
        "refresh-state",
        help="Realign calc state fingerprints with DWD tails (no DWS recalc)",
    )
    _add_refresh_state_args(p)

    p = ops_sp.add_parser("prune", help="Prune superseded DWS snapshots")
    _add_prune_args(p)

    p = ops_sp.add_parser(
        "repair-weekly",
        help="Repair weekly data after date_trunc('week') fix",
    )
    _add_repair_weekly_args(p)

    p = ops_sp.add_parser(
        "spec-status",
        help="Show v_dq_spec_freshness for anchor date(s)",
    )
    p.add_argument("--date", required=True, help="Analysis anchor YYYYMMDD")


def cmd_ops_spec_status(args):
    from backend.db.connection import get_connection
    from backend.etl.ops_spec_status import cmd_spec_status

    con = get_connection()
    try:
        trade_date = _ensure_trade_date(con, _resolve_trade_date(con, args.date))
        cmd_spec_status(con, trade_date)
    finally:
        con.close()


def main():
    from backend.cli_dates import add_date_range_arguments

    p = argparse.ArgumentParser(prog="tradeanalysis")
    sp = p.add_subparsers(dest="command")

    sp.add_parser("check", help="Check environment connectivity")

    # fetch
    fp = sp.add_parser("fetch", help="Pull ODS data into DuckDB")
    fp.add_argument("--ts-code", nargs="+",
                    help="Stock codes to fetch (omitted = all stocks)")
    fp.add_argument("--start", help="Start date YYYYMMDD (default 20150101)")
    fp.add_argument("--end", help="End date YYYYMMDD (default today)")

    # calc
    cp = sp.add_parser(
        "calc",
        help="Compute DWS indicators (calc_date must be <= ODS max; CALC_STRICT_DATE=0 caps)",
    )
    cp.add_argument(
        "--date",
        help="Analysis date YYYYMMDD (default: today; rejected if ahead of ODS max)",
    )
    cp.add_argument("--ts-code", nargs="+",
                    help="Stock codes to calculate (omitted = all stocks)")
    cp.add_argument("--force", action="store_true",
                    help="Recalculate even if calc_date already completed")
    cp.add_argument(
        "--refresh-spec",
        metavar="INDICATORS",
        help="Narrow FULL for stale spec_version only (e.g. ma or ma,volume); skips run_calc",
    )
    cp.add_argument(
        "--dry-run",
        action="store_true",
        help="With --refresh-spec: report stale groups only; no DWS writes",
    )

    # export
    xp = sp.add_parser("export", help="Export analysis wide table to Excel")
    add_date_range_arguments(xp)
    xp.add_argument("--output", default=None,
                    help="Output Excel path (single-day only; range uses per-day default)")
    xp.add_argument("--ts-code", nargs="+", help="Stock codes to export")
    xp.add_argument("--db-path")
    xp.add_argument("--include-st", action="store_true")
    xp.add_argument("--no-index", action="store_true")
    xp.add_argument("--portfolio-file", default=None,
                    help="Path to portfolio stock list xlsx (default: 持仓股列表.xlsx in cwd)")

    # query
    qp = sp.add_parser("query", help="Query DWS indicators")
    qp.add_argument("--ts-code", required=True)
    qp.add_argument("--freq", default="daily")

    # run
    rp = sp.add_parser("run", help="One-command daily analysis: fetch → calc → export")
    add_date_range_arguments(rp)
    rp.add_argument("--ts-code", nargs="+", help="Stock codes (omitted = all stocks)")
    rp.add_argument("--output", default=None,
                    help="Output Excel path. Default: exports/analysis_{date}_gen{now}.xlsx")
    rp.add_argument("--db-path", help="DuckDB file path (default: data/tradeanalysis.duckdb)")
    rp.add_argument("--include-st", action="store_true")
    rp.add_argument("--no-index", action="store_true")
    rp.add_argument("--force", action="store_true",
                    help="Force recalc even if calc_date already completed")
    rp.add_argument("--skip-export", action="store_true",
                    help="Skip Excel export (same-day rerun when report unchanged)")
    rp.add_argument("--continue-on-error", action="store_true",
                    help="With --from/--to: continue after a failed day (default: fail-fast)")
    rp.add_argument("--portfolio-file", default=None,
                    help="Path to portfolio stock list xlsx (default: 持仓股列表.xlsx in cwd)")

    # refresh
    rfp = sp.add_parser(
        "refresh",
        help="Force recalc: fetch → DWD → FULL calc (R1); bypasses run idempotent skip",
    )
    add_date_range_arguments(rfp)
    rfp.add_argument("--ts-code", nargs="+", help="Stock codes (default: all active)")
    rfp.add_argument(
        "--indicator",
        help="Comma-separated indicators (macd,ma,...); default = all 12 routes",
    )
    rfp.add_argument("--export", action="store_true", help="Export Excel after calc")
    rfp.add_argument("--dry-run", action="store_true",
                     help="Print scope estimate only; no API or writes")
    rfp.add_argument("--confirm", action="store_true",
                     help="Confirm large-scope refresh (required above route threshold)")
    rfp.add_argument("--output", default=None, help="Export path when --export")
    rfp.add_argument("--db-path", help="DuckDB file path")
    rfp.add_argument("--continue-on-error", action="store_true",
                     help="With --from/--to: continue after a failed day")

    # ops (maintenance)
    ops_p = sp.add_parser("ops", help="Maintenance / backfill commands")
    ops_sp = ops_p.add_subparsers(dest="ops_command")
    _register_ops_subparsers(ops_sp)

    # backfill-state (top-level alias — deprecated)
    bsp = sp.add_parser(
        "backfill-state",
        help="Backfill dws_calc_state for stocks missing append-routing state",
    )
    _add_backfill_state_args(bsp)

    # backfill-dde-meta (top-level alias — deprecated)
    bdm = sp.add_parser(
        "backfill-dde-meta",
        help="Backfill net_amount_dc + circ_mv for B4 weekly DDE trend (ops)",
    )
    _add_backfill_dde_meta_args(bdm)

    # repair-dde-trend (top-level alias — deprecated)
    rdt = sp.add_parser(
        "repair-dde-trend",
        help="Invalidate DDE trend routing+DWS and CALC_FORCE_HARD recalc",
    )
    _add_repair_dde_trend_args(rdt)

    # prune (top-level alias — deprecated)
    pp = sp.add_parser("prune", help="Prune superseded DWS snapshots (keep last N runs)")
    _add_prune_args(pp)

    # repair-weekly (top-level alias — deprecated)
    rwp = sp.add_parser("repair-weekly",
                        help="Repair weekly data after date_trunc('week') fix (dry-run default)")
    _add_repair_weekly_args(rwp)

    # refresh-state (top-level alias — deprecated)
    rsp = sp.add_parser(
        "refresh-state",
        help="Realign calc state fingerprints with DWD tails (no DWS recalc)",
    )
    _add_refresh_state_args(rsp)

    sp.add_parser("status", help="Show database table stats")

    args = p.parse_args()

    # Assign a unique run ID for this CLI invocation
    if args.command:
        set_run_id(uuid.uuid4().hex[:8])

    handlers = {
        "check": cmd_check,
        "fetch": cmd_fetch,
        "calc": cmd_calc,
        "export": cmd_export,
        "run": cmd_run,
        "refresh": cmd_refresh,
        "backfill-state": cmd_backfill_state,
        "backfill-dde-meta": cmd_backfill_dde_meta,
        "repair-dde-trend": cmd_repair_dde_trend,
        "refresh-state": cmd_refresh_state,
        "prune": cmd_prune,
        "repair-weekly": cmd_repair_weekly,
        "spec-status": cmd_ops_spec_status,
        "query": cmd_query,
        "status": cmd_status,
    }

    if args.command == "ops":
        handler = handlers.get(args.ops_command)
        if handler:
            handler(args)
        else:
            p.print_help()
        return

    if args.command in DEPRECATED_OPS_COMMANDS:
        _warn_deprecated_top_level(args.command)

    handler = handlers.get(args.command)
    if handler:
        handler(args)
    else:
        p.print_help()


if __name__ == "__main__":
    main()
