"""Export index analysis to Excel — standalone sheet matching stock sheet styling."""
import logging

import pandas as pd

logger = logging.getLogger(__name__)

# View for index daily data
_INDEX_VIEW = "v_ads_market_index_daily"

# Borrow Chinese column names from stock sheet + add index-specific entries
from backend.export_wide import _COL_NAMES as _STOCK_COL_NAMES

_INDEX_COL_NAMES = {
    **_STOCK_COL_NAMES,
    "ts_code": "指数代码",
    "index_name": "指数名称",
    "index_category": "分类",
    "pb": "市净率",
}

# Header styling constants — aligned with backend/export_wide.py
_HEADER_FONT_11 = {"name": "微软雅黑", "color": "FFFFFF", "size": 11, "bold": True}
_HEADER_FONT_10 = {"name": "微软雅黑", "color": "FFFFFF", "size": 10}
_BASIC_FILL = "1A1A1A"
_GROUP_FILL = "1A5276"

# Indicator group colors — same palette as export_wide.py _GROUP_COLORS
_COL_TINT = {
    "macd_": "8E44AD", "dif": "8E44AD", "dea": "8E44AD",
    "ma_": "2980B9", "bias_": "2980B9", "ma5": "2980B9", "ma10": "2980B9",
    "vol_": "27AE60", "volume_": "27AE60", "pct_vol": "27AE60",
}
_DEFAULT_TINT = "7F8C8D"

# Column order + group labels (Row 1 merge spans)
_COL_GROUPS = [
    ("基本信息", [
        "ts_code", "index_name", "index_category",
        "close", "pct_chg", "vol", "amount",
        "pe_ttm", "pb", "total_mv",
    ]),
    ("MACD 指标", [
        "dif", "dea", "macd_bar", "macd_zone", "macd_turning_point",
        "macd_trend", "macd_trend_strength", "macd_divergence",
    ]),
    ("MA 均线", [
        "ma_5", "ma_10", "bias_ma5", "ma5_slope", "ma_alignment",
    ]),
    ("量能信号", [
        "volume_ratio", "vol_trend_strength", "vol_zone", "vol_trend",
    ]),
]


def _chinese_name(col: str) -> str:
    """Map English column key to Chinese display name."""
    return _INDEX_COL_NAMES.get(col, col)


def _tint_for_col(col: str) -> str:
    """Match column to indicator color prefix."""
    for prefix, color in _COL_TINT.items():
        if col.startswith(prefix):
            return color
    return _DEFAULT_TINT


def export_index_sheet(con, trade_date: str, ws) -> int:
    """Write index overview to an openpyxl worksheet. Returns row count.

    Styling matches the stock analysis sheets:
    - Row 1: merged group headers (blue fill, white 雅黑 11pt)
    - Row 2: individual **Chinese** column names with indicator-colored fills
             + hover comments from export-column-comments.yaml
    - Data from Row 3
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from backend.export_wide import _attach_header_comment

    df = con.execute(f"""
        SELECT * FROM {_INDEX_VIEW}
        WHERE trade_date = ?
        ORDER BY
            CASE
                WHEN ts_code = '000001.SH' THEN 0
                WHEN ts_code LIKE '%.SH' THEN 1
                WHEN ts_code LIKE '%.SZ' THEN 2
                ELSE 3
            END,
            ts_code
    """, [trade_date]).df()

    if df.empty:
        ws.cell(row=1, column=1, value=f"No index data for {trade_date}")
        return 0

    # Flatten column order from groups
    all_cols = []
    for _, cols in _COL_GROUPS:
        for c in cols:
            if c in df.columns and c not in all_cols:
                all_cols.append(c)
    df = df[all_cols]

    # ── styles ────────────────────────────────────────────
    group_font = Font(**_HEADER_FONT_11)
    col_font = Font(**_HEADER_FONT_10)
    group_fill = PatternFill(start_color=_GROUP_FILL, end_color=_GROUP_FILL, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    header_border = Border(bottom=Side(style="thin", color="5D6D7E"))

    # ── Row 1: merged group headers ───────────────────────
    col = 1
    for group_name, group_cols in _COL_GROUPS:
        present = [c for c in group_cols if c in all_cols]
        if not present:
            continue
        n = len(present)
        if n > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + n - 1)
        c = ws.cell(row=1, column=col, value=group_name)
        c.fill = group_fill
        c.font = group_font
        c.alignment = center_align
        c.border = header_border
        col += n

    # ── Row 2: individual column names (Chinese + comments) ─
    for j, eng_name in enumerate(all_cols, 1):
        cn_name = _chinese_name(eng_name)
        tint = _tint_for_col(eng_name)
        cell = ws.cell(row=2, column=j, value=cn_name)
        cell.font = col_font
        cell.fill = PatternFill(start_color=tint, end_color=tint, fill_type="solid")
        cell.alignment = center_align
        cell.border = header_border
        _attach_header_comment(cell, eng_name, weekly=False)

    # ── Data rows ─────────────────────────────────────────
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col_name in enumerate(all_cols, 1):
            val = row[col_name]
            if isinstance(val, float) and pd.isna(val):
                val = None
            ws.cell(row=i + 3, column=j, value=val)

    # ── Column widths (based on Chinese header) ────────────
    from openpyxl.utils import get_column_letter
    for j, eng_name in enumerate(all_cols, 1):
        cn_name = _chinese_name(eng_name)
        hdr_len = sum(2.2 if "一" <= c <= "鿿" else 1.0 for c in cn_name)
        ws.column_dimensions[get_column_letter(j)].width = min(hdr_len + 3, 22)

    ws.auto_filter.ref = ws.dimensions
    return len(df)
