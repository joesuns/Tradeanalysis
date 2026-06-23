import logging
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment

from backend.etl.divergence_tradable import TradableEnrichStats, enrich_tradable_columns
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

VIEW_MAP = {
    "daily": "v_ads_analysis_wide_daily",
    "weekly": "v_ads_analysis_wide_weekly",
}
INDEX_VIEW_MAP = {
    "daily": "v_ads_index_wide",
    "weekly": "v_ads_index_wide_weekly",
}

EXPORT_DIR = "exports"


@dataclass
class ExportResult:
    """Export outcome with tradable enrich observability."""

    row_count: int
    tradable_enrich: Dict[str, dict]

    def __int__(self) -> int:
        return self.row_count


def format_tradable_enrich_log(stats: TradableEnrichStats) -> str:
    return (
        f"progress export: tradable enrich {stats.freq} | "
        f"l1_macd={stats.l1_macd} l1_dde={stats.l1_dde} "
        f"tradable={stats.tradable} reject={stats.reject} | "
        f"{stats.elapsed_sec:.1f}s"
    )


def log_tradable_enrich_progress(stats: TradableEnrichStats) -> None:
    logger.info(format_tradable_enrich_log(stats))


def build_export_data_completeness(
    analysis_date: str,
    tradable_enrich: Dict[str, dict],
) -> dict:
    return {
        "analysis_date": analysis_date,
        "tradable_enrich": tradable_enrich,
    }


def default_export_path(trade_date: str, output: str = None) -> str:
    """Default Excel path under exports/; explicit output is returned unchanged."""
    if output is not None:
        return output
    gen_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{EXPORT_DIR}/analysis_{trade_date}_gen{gen_ts}.xlsx"


# Column name translations (English → Chinese) — with units where applicable
_COL_NAMES = {
    "freq": "周期", "trade_date": "交易日期", "ts_code": "股票代码",
    "stock_code": "代码", "stock_name": "股票名称", "exchange": "交易所",
    "sector": "上市板", "industry": "申万行业",
    "tdx_industry_board": "通达信行业", "dc_concept_board": "概念板块",
    "dc_theme_board": "所属题材", "is_st": "ST",
    "close": "收盘价", "pct_chg": "涨跌幅%",
    "pct_chg_3d": "最近3天涨跌幅", "pct_chg_1m": "最近1月涨跌幅", "pct_chg_1y": "最近1年涨跌幅",
    "vol": "成交量(万手)", "amount": "成交额(亿)",
    "total_mv": "总市值(亿)", "pe_ttm": "市盈率", "turnover_rate": "换手率%",
    "kpattern": "K线形态", "kpattern_strength": "形态强度",
    "ema_12": "EMA12", "ema_26": "EMA26", "dif": "DIF", "dea": "DEA",
    "macd_bar": "MACD柱",
    "macd_divergence": "MACD结构背离",
    "macd_divergence_tradable": "MACD可交易背离",
    "macd_divergence_reject": "MACD背离剔除",
    "macd_zone": "MACD区域",
    "macd_turning_point": "MACD转折", "macd_alert": "MACD警惕", "macd_trend": "MACD趋势",
    "macd_trend_strength": "MACD趋势强度",
    "ma_5": "MA5", "ma_10": "MA10",
    "bias_ma5": "MA5乖离率", "bias_ma10": "MA10乖离率",
    "ma5_slope": "MA5斜率", "ma10_slope": "MA10斜率",
    "ma_alignment": "均线形态", "ma_turning_point": "均线转折",
    "net_mf_amount": "主力净流入(万元)", "ddx": "DDX", "ddx2": "DDX2",
    "dde_trend": "DDE趋势", "dde_trend_strength": "DDE趋势强度", "dde_alert": "DDE警惕",
    "dde_divergence": "DDE结构背离",
    "dde_divergence_tradable": "DDE可交易背离",
    "dde_divergence_reject": "DDE背离剔除",
    "ma_vol_5": "5日均量(万手)", "pct_vol_rank": "量能百分位",
    "vol_zone": "量能区域", "vol_trend": "量能趋势",
    "volume_ratio": "量比", "vol_ratio": "量比", "vol_trend_strength": "量能趋势强度",
    "vol_divergence": "量价背离",
    "price_position_60d": "60日价格滚动分位(%)", "price_position_120d": "120日价格滚动分位(%)",
    "price_position_250d": "250日价格滚动分位(%)",
    "vol_signal": "量价信号",
    "risk_alert": "风险预警",
}

_BASIC_HEADER_FILL = "1A1A1A"

_ID_COLS = [
    "ts_code", "trade_date", "stock_code", "stock_name",
    "exchange", "sector", "industry", "tdx_industry_board", "dc_concept_board",
    "dc_theme_board", "is_st",
]
_FUND_COLS = [
    "close", "pct_chg", "pct_chg_3d", "pct_chg_1m", "pct_chg_1y",
    "vol", "amount", "total_mv",
    "pe_ttm", "turnover_rate", "volume_ratio",
]

_GROUP_COLORS = {
    "kpattern": "C0392B",
    "price_position_": "E74C3C",
    "ema_": "8E44AD", "macd_": "8E44AD", "dif": "8E44AD", "dea": "8E44AD",
    "ma_vol_": "27AE60",
    "ma_": "2980B9", "bias_": "2980B9", "ma5_": "2980B9", "ma10_": "2980B9",
    "dde_": "D35400", "ddx": "D35400", "net_mf": "D35400",
    "vol_": "27AE60", "pct_vol": "27AE60",
}
_DEFAULT_GROUP_COLOR = "7F8C8D"

# 综合分析 sheet：可交易背离 + 量价信号；不含 MACD/DDE 结构/剔除背离
_SIGNAL_ONLY = [
    "kpattern", "kpattern_strength",
    "price_position_60d", "price_position_120d", "price_position_250d",
    "macd_divergence_tradable",
    "macd_zone", "macd_turning_point", "macd_alert", "macd_trend", "macd_trend_strength",
    "ma_alignment", "ma_turning_point", "bias_ma5", "bias_ma10",
    "dde_divergence_tradable",
    "dde_trend", "dde_trend_strength", "dde_alert", "risk_alert",
    "vol_zone", "vol_trend", "vol_divergence", "vol_signal",
]


def _resolve_portfolio_remarks(
    con,
    portfolio_codes: "List[str]",
    trade_date: str,
    daily_stock_codes: "set",
) -> "Dict[str, str]":
    """Resolve remark for each portfolio stock based on data presence.

    Returns:
        Dict mapping stock_code → remark label:
        - ``"正常"`` — stock has daily data for trade_date
        - ``"已退市"`` — stock in dim_stock with delist_date < trade_date
        - ``"当日无数据（可能停牌）"`` — in dim_stock, not delisted, no daily data
        - ``"未入库"`` — not in dim_stock at all
    """
    if not portfolio_codes:
        return {}

    placeholders = ",".join(["?" for _ in portfolio_codes])
    dim_rows = con.execute(
        f"SELECT stock_code, delist_date FROM dim_stock WHERE stock_code IN ({placeholders})",
        portfolio_codes,
    ).fetchall()

    dim_map: dict = {}
    for stock_code, delist_date in dim_rows:
        dim_map[stock_code] = delist_date

    result = {}
    for code in portfolio_codes:
        if code in daily_stock_codes:
            result[code] = "正常"
        elif code not in dim_map:
            result[code] = "未入库"
        elif dim_map[code] and str(dim_map[code]) < trade_date:
            result[code] = "已退市"
        else:
            result[code] = "当日无数据（可能停牌）"

    return result


def _reorder_vol_signal(cols):
    """Place vol_signal adjacent to vol_divergence (fallback: after vol_trend)."""
    if "vol_signal" not in cols:
        return cols
    out = list(cols)
    out.remove("vol_signal")
    if "vol_divergence" in out:
        out.insert(out.index("vol_divergence") + 1, "vol_signal")
    elif "vol_trend" in out:
        out.insert(out.index("vol_trend") + 1, "vol_signal")
    else:
        out.append("vol_signal")
    return out


def _attach_header_comment(cell, col_key, weekly=False):
    """Attach YAML glossary comment to a header cell (row 2)."""
    from backend.export_column_comments import (
        comment_author,
        comment_box_size,
        format_column_comment,
    )

    text = format_column_comment(col_key, weekly=weekly)
    if not text:
        return
    width, height = comment_box_size()
    cell.comment = Comment(text, comment_author(), width=width, height=height)

# Shared alert enum labels — macd_alert and dde_alert use identical translations.
_ALERT_LABELS = {
    "upturn_reverse": "上升趋势回落",
    "downturn_reverse": "下降趋势反弹",
    "upturn_flat": "上升趋势走平",
    "downturn_flat": "下降趋势走平",
}

# Enum value translations (English → Chinese). NULL = no signal, shown as "-"
_ENUM_VALUES = {
    "kpattern": {"yang_bao_yin": "阳包阴", "yang_ke_yin": "阳克阴",
                 "yin_bao_yang": "阴包阳", "yin_ke_yang": "阴克阳",
                 "contrarian_yin_bao_yang": "阴包阳(反向买入)",
                 "contrarian_yin_ke_yang": "阴克阳(反向买入)",
                 "mu_bei_xian": "墓碑线", "bi_lei_zhen": "避雷针",
                 "gao_kai_chang_yin": "高开长阴"},
    "macd_zone": {"bull": "多头", "bear": "空头"},
    "macd_trend": {"up": "上升", "down": "下降", "flat": "走平"},
    "macd_turning_point": {"golden_cross": "金叉", "dead_cross": "死叉",
                           "near_golden": "近金叉", "near_dead": "近死叉"},
    "macd_divergence": {"top_divergence": "顶背离", "bottom_divergence": "底背离"},
    "macd_divergence_tradable": {"top_divergence": "顶背离", "bottom_divergence": "底背离"},
    "macd_divergence_reject": {"skip_peak": "隔峰", "tg_lag": "滞后", "zone_mismatch": "区域"},
    "macd_alert": _ALERT_LABELS,
    "ma_turning_point": {"golden_cross": "金叉", "dead_cross": "死叉",
                         "near_golden": "近金叉", "near_dead": "近死叉"},
    "dde_trend": {"up": "上升", "down": "下降", "flat": "走平"},
    "dde_divergence": {"top_divergence": "顶背离", "bottom_divergence": "底背离"},
    "dde_divergence_tradable": {"top_divergence": "顶背离", "bottom_divergence": "底背离"},
    "dde_divergence_reject": {"skip_peak": "隔峰", "tg_lag": "滞后", "zone_mismatch": "区域"},
    "dde_alert": _ALERT_LABELS,
    "vol_zone": {"explosive": "爆量", "low_volume": "地量", "normal": "正常"},
    "vol_trend": {"expanding": "放量", "shrinking": "缩量", "flat": "平量"},
    "vol_divergence": {"top_divergence": "顶背离", "bottom_divergence": "底背离"},
    "vol_signal": {
        "breakout_tight": "缩量突破", "breakout_moderate": "温和突破",
        "breakout_heavy": "爆量突破",
        "volume_climax": "放量滞涨",
        "volume_dry_up": "缩量止跌",
        "golden_cross_weakened": "金叉量弱", "dead_cross_weakened": "死叉量弱",
    },
}

# Event columns — NULL means "no signal today" (shown as "-")
_EVENT_SIGNAL_COLS = {
    "kpattern", "kpattern_strength",
    "macd_divergence", "macd_divergence_tradable", "macd_divergence_reject",
    "macd_turning_point", "macd_alert",
    "ma_turning_point", "dde_alert",
    "dde_divergence", "dde_divergence_tradable", "dde_divergence_reject",
    "vol_divergence", "vol_signal", "risk_alert",
}
# State metrics — NULL means "not computable / insufficient history" (shown as "N/A")
_STATE_METRIC_COLS = {
    "pct_vol_rank", "vol_zone", "ma_alignment",
    "macd_zone", "macd_trend", "macd_trend_strength",
    "dde_trend", "dde_trend_strength",
    "vol_trend", "vol_trend_strength", "volume_ratio",
    "price_position_60d", "price_position_120d", "price_position_250d",
}
_FUNDAMENTAL_NA_COLS = {"pe_ttm", "pct_chg_3d", "pct_chg_1m", "pct_chg_1y"}
# Classification columns — stock attributes not derived from signals; null → "N/A"
_PLATE_CLASSIFICATION_COLS = {"tdx_industry_board", "dc_concept_board", "dc_theme_board"}
# Backward-compatible union for highlight logic
_SIGNAL_COLS = _EVENT_SIGNAL_COLS | _STATE_METRIC_COLS | _FUNDAMENTAL_NA_COLS

# Columns to round to 2 decimal places
_ROUND_2DP = {"close", "pct_chg", "pct_chg_3d", "pct_chg_1m", "pct_chg_1y",
              "pe_ttm", "turnover_rate", "net_mf_amount",
              "volume_ratio", "vol_trend_strength",
              "price_position_60d", "price_position_120d", "price_position_250d"}

# Columns to convert 万元 → 亿 (divide by 10000)
_CONVERT_DIV10000 = {"total_mv", "vol", "ma_vol_5"}  # → 万 (or 亿 for mv)
_CONVERT_DIV10 = set()  # unused, kept for clarity
_CONVERT_AMOUNT = {"amount"}  # 千元 → 亿 (/100000)

# Weekly column name overrides
_WEEKLY_OVERRIDE = {"ma_vol_5": "5周均量(万手)"}


def export_wide_to_excel(
    db_path: str,
    trade_date: str,       # YYYYMMDD
    output_path: str = "",  # .xlsx path (auto-timestamped if empty)
    filter_st: bool = True,
    include_index: bool = True,
    ts_codes: list[str] = None,  # 可选，只导出指定股票
    portfolio_stocks: list[dict] = None,  # [{"stockcode","stockname"}, ...]
) -> ExportResult:
    """Export horizontal daily+weekly merged analysis to Excel.

    Each row = one stock on trade_date, with daily indicators on the left
    and weekly (week-to-date) indicators on the right. Two-row header:
    Row 1 merges group labels (日线指标/周线指标), Row 2 has individual column names.
    Returns ExportResult with row_count and tradable_enrich stats.
    """
    tradable_meta: Dict[str, dict] = {}
    logger.info("progress export: started | date=%s", trade_date)
    t0 = time.monotonic()
    con = duckdb.connect(db_path)

    from backend.config import EXPORT_SPEC_GATE
    if EXPORT_SPEC_GATE:
        from backend.etl.calc_spec_gate import export_spec_freshness_warnings

        for msg in export_spec_freshness_warnings(con, trade_date):
            logger.warning("export spec gate: %s", msg)

    # ---- Daily data ----
    daily = con.execute(
        f"SELECT * FROM {VIEW_MAP['daily']} WHERE trade_date = ?"
        + (" AND is_st = 0" if filter_st else ""),
        [trade_date]
    ).df()
    if daily.empty:
        con.close()
        logger.info("progress export: done | rows=0 | %.0fs", time.monotonic() - t0)
        return ExportResult(0, tradable_meta)

    logger.info(
        "progress export: daily query done | rows=%d | %.0fs",
        len(daily), time.monotonic() - t0,
    )

    # ---- Optional ts_code filter ----
    if ts_codes:
        daily = daily[daily["ts_code"].isin(ts_codes)]
        if daily.empty:
            con.close()
            return ExportResult(0, tradable_meta)

    daily = _format_numbers(daily)
    # ── Market summary columns (上证/沪深300) ──
    try:
        idx_summary = con.execute("""
            SELECT ts_code, pct_chg, macd_trend
            FROM v_ads_market_index_daily
            WHERE trade_date = ?
              AND ts_code IN ('000001.SH', '000300.SH')
        """, [trade_date]).df()
        for _, row in idx_summary.iterrows():
            code = row["ts_code"]
            suffix = "000001" if code == "000001.SH" else "000300"
            daily[f"index_{suffix}_pct_chg"] = row["pct_chg"]
            macd_val = row.get("macd_trend")
            daily[f"index_{suffix}_macd"] = macd_val if macd_val and not (isinstance(macd_val, float) and pd.isna(macd_val)) else None
    except Exception:
        pass  # index data not available
    daily, daily_enrich = enrich_tradable_columns(daily, con, freq="daily")
    tradable_meta["daily"] = daily_enrich.to_dict()
    log_tradable_enrich_progress(daily_enrich)

    # ---- Enrich with plate/concept data ----
    from backend.fetch.ods_plate import load_plate_enrichment

    t_plate = time.monotonic()
    logger.info("progress export: loading plate enrichment | date=%s", trade_date)
    plate_enrichment = load_plate_enrichment(con, trade_date)
    if plate_enrichment:
        plate_df_data = []
        for ts_code, cols in plate_enrichment.items():
            plate_df_data.append({
                "ts_code": ts_code,
                "tdx_industry_board": cols.get("tdx_industry_board"),
                "dc_concept_board": cols.get("dc_concept_board"),
                "dc_theme_board": cols.get("dc_theme_board"),
            })
        plate_df = pd.DataFrame(plate_df_data)
        daily = daily.merge(plate_df, on="ts_code", how="left")
        logger.info(
            "progress export: plate enrichment done | enriched=%d rows | %.0fs",
            len(plate_df), time.monotonic() - t_plate,
        )
    else:
        logger.info("progress export: no plate data for %s | %.0fs",
                    trade_date, time.monotonic() - t_plate)

    # Fill missing plate/concept values with "N/A"
    for col in ["tdx_industry_board", "dc_concept_board", "dc_theme_board"]:
        if col in daily.columns:
            daily[col] = daily[col].fillna("N/A")

    # ---- Weekly data: use latest week-end ≤ trade_date ----
    t_weekly = time.monotonic()
    logger.info("progress export: loading weekly | date=%s", trade_date)
    week_end = con.execute(
        "SELECT MAX(trade_date) FROM dim_date "
        "WHERE trade_date <= ? AND is_week_end = 1",
        [trade_date]
    ).fetchone()[0]
    weekly = con.execute(
        f"SELECT * FROM {VIEW_MAP['weekly']} WHERE trade_date = ?"
        + (" AND is_st = 0" if filter_st else ""),
        [week_end]
    ).df() if week_end else pd.DataFrame()
    logger.info(
        "progress export: weekly query done | week_end=%s rows=%d | %.0fs",
        week_end or "-", len(weekly), time.monotonic() - t_weekly,
    )
    if ts_codes:
        weekly = weekly[weekly["ts_code"].isin(ts_codes)]
    weekly = _format_numbers(weekly)
    if not weekly.empty:
        weekly, weekly_enrich = enrich_tradable_columns(weekly, con, freq="weekly")
        tradable_meta["weekly"] = weekly_enrich.to_dict()
        log_tradable_enrich_progress(weekly_enrich)
    else:
        empty_weekly = TradableEnrichStats(freq="weekly")
        tradable_meta["weekly"] = empty_weekly.to_dict()

    # Drop identity + fundamental columns from weekly (already in basic section from daily)
    id_cols_drop = ["freq", "trade_date", "stock_code", "stock_name",
                    "exchange", "sector", "industry", "is_st",
                    "close", "pct_chg", "vol", "amount", "total_mv",
                    "pe_ttm", "turnover_rate", "volume_ratio"]
    # Keep ts_code for merge
    weekly = weekly.drop(columns=[c for c in id_cols_drop if c in weekly.columns], errors="ignore")

    # Track which columns are daily vs weekly
    daily_cols = _reorder_vol_signal([c for c in daily.columns if c != "freq"])
    weekly_cols = list(weekly.columns)

    # Basic info columns (for both sheets)
    basic_cols_outer = _ID_COLS + [c for c in _FUND_COLS if c in daily_cols]

    # Add __w__ prefix to weekly indicator columns (not ts_code — needed for merge)
    weekly_indicator_cols = [c for c in weekly_cols if c != "ts_code"]
    weekly_renamed = weekly.rename(columns={c: f"__w__{c}" for c in weekly_indicator_cols})

    # LEFT JOIN on ts_code. When weekly is empty (e.g. no week-end ≤ trade_date),
    # the frame has no ts_code column — keep daily rows, omit weekly indicators.
    if "ts_code" in weekly_renamed.columns:
        merged = daily.merge(weekly_renamed, on="ts_code", how="left")
        weekly_cols = weekly_indicator_cols  # for header building, exclude ts_code
    else:
        merged = daily.copy()
        weekly_cols = []

    logger.info("progress export: weekly merge done | %.0fs", time.monotonic() - t0)

    # ---- Write to Excel ----
    t_sheets = time.monotonic()
    logger.info("progress export: building sheets | rows=%d", len(merged))
    wb = Workbook()
    wb.remove(wb.active)
    # ---- Signal-only analysis sheet layout (reused by portfolio) ----
    _signal_set = set(_SIGNAL_ONLY)
    daily_signal_only = [c for c in _SIGNAL_ONLY if c in daily_cols] + [
        c for c in daily_cols if c in basic_cols_outer and c not in _signal_set
    ]
    weekly_signal_only = [c for c in _SIGNAL_ONLY if c in weekly_cols]

    display_full, layout_full = _build_merged_display_df(merged, daily_cols, weekly_cols)
    layout_signal = _resolve_sheet_layout(
        daily_signal_only, weekly_signal_only, merged.columns,
    )

    # ---- Portfolio sheet (first sheet, if portfolio_stocks provided) ----
    if portfolio_stocks:
        portfolio_codes = [s["stockcode"] for s in portfolio_stocks]
        portfolio_name_map = {s["stockcode"]: s["stockname"] for s in portfolio_stocks}

        # Filter merged data to portfolio stocks via stock_code column
        if "stock_code" in merged.columns:
            portfolio_df = merged[merged["stock_code"].isin(portfolio_codes)].copy()
        else:
            portfolio_df = merged[merged["ts_code"].isin(portfolio_codes)].copy()

        # Detect which portfolio stocks have data
        daily_stock_codes = set(portfolio_df["stock_code"].tolist()) if "stock_code" in portfolio_df.columns else set()

        # Resolve remarks for ALL portfolio stocks
        remarks = _resolve_portfolio_remarks(
            con, portfolio_codes, trade_date, daily_stock_codes,
        )

        # Build placeholder rows for missing portfolio stocks
        missing_codes = [c for c in portfolio_codes if c not in daily_stock_codes]
        if missing_codes:
            placeholder_rows = []
            for code in missing_codes:
                row = {col: None for col in layout_signal.display_cols}
                row["stock_code"] = code
                row["stock_name"] = portfolio_name_map.get(code, code)
                row["ts_code"] = ""
                placeholder_rows.append(row)
            placeholder_df = pd.DataFrame(placeholder_rows)
            # Ensure all display columns exist
            for col in layout_signal.display_cols:
                if col not in placeholder_df.columns:
                    placeholder_df[col] = None
            placeholder_df = placeholder_df[layout_signal.display_cols]
            # Apply display nulls to placeholder rows
            placeholder_df = apply_display_nulls(placeholder_df)
            # Combine: data rows first, then placeholder rows
            portfolio_display = pd.concat(
                [portfolio_df[layout_signal.display_cols], placeholder_df],
                ignore_index=True,
            )
        else:
            portfolio_display = portfolio_df[layout_signal.display_cols]

        # Apply display transformations to the portfolio subset
        portfolio_display = _transform_display_values(portfolio_display)

        # Append remark column
        if "stock_code" in portfolio_df.columns:
            data_codes = portfolio_df["stock_code"].tolist()
        else:
            data_codes = []
        all_codes = data_codes + missing_codes
        remark_values = [remarks.get(c, "") for c in all_codes]
        portfolio_display["备注"] = remark_values

        # Build layout for portfolio (same as signal, "备注" as trailing non-signal column)
        portfolio_layout = SheetLayout(
            display_cols=layout_signal.display_cols + ["备注"],
            chinese_names=layout_signal.chinese_names + ["备注"],
            basic_cols=layout_signal.basic_cols,
            daily_signal=layout_signal.daily_signal,
            weekly_signal=layout_signal.weekly_signal,
            basic_names=layout_signal.basic_names,
            daily_names=layout_signal.daily_names,
            weekly_names=layout_signal.weekly_names,
            n_basic=layout_signal.n_basic,
            n_daily=layout_signal.n_daily,
            n_weekly=layout_signal.n_weekly,
        )

        _write_sheet_from_display(wb, "持仓股分析", portfolio_display, portfolio_layout)

        # Patch header for remark column (last column, merged rows 1-2 like basic cols)
        ws = wb["持仓股分析"]
        remark_col = layout_signal.n_basic + layout_signal.n_daily + layout_signal.n_weekly + 1
        ws.merge_cells(start_row=1, start_column=remark_col, end_row=2, end_column=remark_col)
        remark_cell = ws.cell(row=1, column=remark_col, value="备注")
        remark_cell.font = Font(name="微软雅黑", color="FFFFFF", size=10)
        remark_cell.fill = PatternFill(start_color=_BASIC_HEADER_FILL, end_color=_BASIC_HEADER_FILL, fill_type="solid")
        remark_cell.alignment = Alignment(horizontal="center", vertical="center")
        remark_cell.border = Border(bottom=Side(style="thin", color="5D6D7E"))
        for r in (1, 2):
            c2 = ws.cell(row=r, column=remark_col)
            c2.fill = PatternFill(start_color=_BASIC_HEADER_FILL, end_color=_BASIC_HEADER_FILL, fill_type="solid")
            c2.font = Font(name="微软雅黑", color="FFFFFF", size=10)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = Border(bottom=Side(style="thin", color="5D6D7E"))

    # ---- Signal-only analysis sheet ----
    display_signal = display_full[layout_signal.display_cols]
    _write_sheet_from_display(wb, "综合分析", display_signal, layout_signal)
    _write_sheet_from_display(wb, "个股分析", display_full, layout_full)

    # ── Index overview sheet ──
    t_idx = time.monotonic()
    logger.info("progress export: building index sheet | date=%s", trade_date)
    try:
        from backend.export_index import export_index_sheet
        ws_index = wb.create_sheet("指数概览")
        n_idx = export_index_sheet(con, trade_date, ws_index)
        logger.info("progress export: index sheet done | rows=%d | %.0fs",
                    n_idx, time.monotonic() - t_idx)
    except Exception as e:
        logger.warning("progress export: index sheet skipped: %s", e)

    # ---- SH Index ----
    if include_index:
        idx_daily = con.execute(
            f"SELECT * FROM {INDEX_VIEW_MAP['daily']} WHERE trade_date = ?", [trade_date]
        ).df()
        idx_daily = _format_numbers(idx_daily)
        if not idx_daily.empty:
            idx_daily = idx_daily.drop(columns=[c for c in id_cols_drop if c in idx_daily.columns], errors="ignore")
            _write_sheet_merged(wb, "上证指数", idx_daily, list(idx_daily.columns), [])

    con.close()
    logger.info("progress export: sheets built | %.0fs", time.monotonic() - t_sheets)

    import os
    if not output_path or output_path == "analysis.xlsx":
        output_path = default_export_path(trade_date)
    parent = os.path.dirname(os.path.abspath(output_path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    t_save = time.monotonic()
    logger.info("progress export: writing xlsx | path=%s", output_path)
    wb.save(output_path)
    logger.info("progress export: xlsx saved | %.0fs", time.monotonic() - t_save)
    logger.info(
        "progress export: done | rows=%d | %.0fs",
        len(merged), time.monotonic() - t0,
    )
    return ExportResult(len(merged), tradable_meta)


def _format_numbers(df: "pd.DataFrame") -> "pd.DataFrame":
    """Round numeric columns + convert 万元→亿 for market cap."""
    for col in _ROUND_2DP:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round(float(x), 2) if pd.notna(x) else x)
    for col in _CONVERT_DIV10000:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round(float(x) / 10000, 2) if pd.notna(x) else x)
    for col in _CONVERT_DIV10:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round(float(x) / 10, 2) if pd.notna(x) else x)
    for col in _CONVERT_AMOUNT:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round(float(x) / 100000, 2) if pd.notna(x) else x)
    return df


def apply_display_nulls(df: pd.DataFrame) -> pd.DataFrame:
    """Apply export display semantics before Chinese rename."""
    df = df.copy()
    for col in _EVENT_SIGNAL_COLS:
        if col in df.columns:
            df[col] = df[col].fillna("-")
        wcol = f"__w__{col}"
        if wcol in df.columns:
            df[wcol] = df[wcol].fillna("-")
    for col in _STATE_METRIC_COLS | _FUNDAMENTAL_NA_COLS:
        if col in df.columns:
            df[col] = df[col].fillna("N/A")
        wcol = f"__w__{col}"
        if wcol in df.columns:
            df[wcol] = df[wcol].fillna("N/A")
    return df


def _translate_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """Translate column names to Chinese, enum values to Chinese, apply null semantics."""
    df = apply_display_nulls(df)
    df = df.rename(columns={c: _COL_NAMES.get(c, c) for c in df.columns})
    for col, mapping in _ENUM_VALUES.items():
        cn_col = _COL_NAMES.get(col, col)
        if cn_col in df.columns:
            df[cn_col] = df[cn_col].map(lambda x: mapping.get(x, x) if pd.notna(x) else x)
    return df


def _chinese_name_for_col(col: str) -> str:
    if col.startswith("__w__"):
        eng = col[5:]
        return _WEEKLY_OVERRIDE.get(eng, _COL_NAMES.get(eng, eng))
    return _COL_NAMES.get(col, col)


def _color_for_indicator(col_eng: str) -> str:
    for prefix, color in _GROUP_COLORS.items():
        if col_eng.startswith(prefix):
            return color
    return _DEFAULT_GROUP_COLOR


@dataclass
class SheetLayout:
    """Resolved column layout for merged daily+weekly export sheets."""

    display_cols: List[str]
    chinese_names: List[str]
    basic_cols: List[str]
    daily_signal: List[str]
    weekly_signal: List[str]
    basic_names: List[str]
    daily_names: List[str]
    weekly_names: List[str]
    n_basic: int
    n_daily: int
    n_weekly: int


def _resolve_sheet_layout(
    daily_cols: List[str],
    weekly_cols: List[str],
    available_columns,
) -> SheetLayout:
    avail = set(available_columns)
    basic_cols = [
        c for c in _ID_COLS + _FUND_COLS
        if c in daily_cols and c in avail
    ]
    basic_set = set(basic_cols)
    daily_signal = [
        c for c in daily_cols
        if c not in basic_set and c != "freq" and c in avail
    ]
    weekly_signal = [
        c for c in weekly_cols if f"__w__{c}" in avail
    ]
    display_cols = (
        basic_cols
        + daily_signal
        + [f"__w__{c}" for c in weekly_signal]
    )
    basic_names = [_COL_NAMES.get(c, c) for c in basic_cols]
    daily_names = [_COL_NAMES.get(c, c) for c in daily_signal]
    weekly_names = [_chinese_name_for_col(f"__w__{c}") for c in weekly_signal]
    chinese_names = [_chinese_name_for_col(c) for c in display_cols]
    return SheetLayout(
        display_cols=display_cols,
        chinese_names=chinese_names,
        basic_cols=basic_cols,
        daily_signal=daily_signal,
        weekly_signal=weekly_signal,
        basic_names=basic_names,
        daily_names=daily_names,
        weekly_names=weekly_names,
        n_basic=len(basic_cols),
        n_daily=len(daily_signal),
        n_weekly=len(weekly_signal),
    )


def _transform_display_values(df: pd.DataFrame) -> pd.DataFrame:
    """Enum translation + export null semantics on English/__w__ columns."""
    out = df.copy()
    for col, mapping in _ENUM_VALUES.items():
        if col in out.columns:
            out[col] = out[col].replace(mapping)
        wcol = f"__w__{col}"
        if wcol in out.columns:
            out[wcol] = out[wcol].replace(mapping)
    return apply_display_nulls(out)


def _build_merged_display_df(
    source: pd.DataFrame,
    daily_cols: List[str],
    weekly_cols: List[str],
) -> Tuple[pd.DataFrame, SheetLayout]:
    layout = _resolve_sheet_layout(daily_cols, weekly_cols, source.columns)
    english = source[layout.display_cols]
    display = _transform_display_values(english)
    return display, layout


def _set_column_widths(ws, display_df: pd.DataFrame, layout: SheetLayout) -> None:
    for col_idx, col_eng in enumerate(layout.display_cols, 1):
        col_name = layout.chinese_names[col_idx - 1]
        header_len = sum(2.2 if "一" <= c <= "鿿" else 1.0 for c in str(col_name))
        width = max(header_len + 2, 8)
        for val in display_df.iloc[:20, col_idx - 1]:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            val_len = sum(2.2 if "一" <= c <= "鿿" else 1.0 for c in str(val))
            width = max(width, min(val_len + 2, 30))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 22)


def _write_sheet_headers(ws, layout: SheetLayout) -> None:
    group_font = Font(name="微软雅黑", color="FFFFFF", size=11)
    group_fill_daily = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    group_fill_weekly = PatternFill(start_color="0D6B6B", end_color="0D6B6B", fill_type="solid")
    col_font = Font(name="微软雅黑", color="FFFFFF", size=10)
    basic_fill = PatternFill(
        start_color=_BASIC_HEADER_FILL, end_color=_BASIC_HEADER_FILL, fill_type="solid",
    )
    header_bottom = Border(bottom=Side(style="thin", color="5D6D7E"))

    n_basic = layout.n_basic
    n_daily = layout.n_daily
    n_weekly = layout.n_weekly
    daily_start = n_basic + 1
    weekly_start = n_basic + n_daily + 1
    weekly_end = n_basic + n_daily + n_weekly

    if n_daily > 0:
        ws.merge_cells(
            start_row=1, start_column=daily_start,
            end_row=1, end_column=daily_start + n_daily - 1,
        )
        c = ws.cell(row=1, column=daily_start, value="日 线 指 标")
        c.fill = group_fill_daily
        c.font = group_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_bottom

    if n_weekly > 0:
        ws.merge_cells(
            start_row=1, start_column=weekly_start,
            end_row=1, end_column=weekly_end,
        )
        c = ws.cell(row=1, column=weekly_start, value="周 线 指 标")
        c.fill = group_fill_weekly
        c.font = group_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_bottom

    for i in range(1, n_basic + 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        ws.cell(row=1, column=i, value=layout.basic_names[i - 1])
        for r in (1, 2):
            c2 = ws.cell(row=r, column=i)
            c2.fill = basic_fill
            c2.font = col_font
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = header_bottom

    for i, name in enumerate(layout.daily_names):
        c = daily_start + i
        eng = layout.daily_signal[i]
        tint = _color_for_indicator(eng)
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = col_font
        cell.fill = PatternFill(start_color=tint, end_color=tint, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_bottom
        _attach_header_comment(cell, eng, weekly=False)

    for i, name in enumerate(layout.weekly_names):
        c = weekly_start + i
        eng = layout.weekly_signal[i]
        tint = _color_for_indicator(eng)
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = col_font
        cell.fill = PatternFill(start_color=tint, end_color=tint, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_bottom
        _attach_header_comment(cell, eng, weekly=True)


def _write_sheet_from_display(
    wb,
    sheet_name: str,
    display_df: pd.DataFrame,
    layout: SheetLayout,
) -> None:
    """Write pre-transformed display DataFrame with two-row header."""
    ws = wb.create_sheet(title=sheet_name)
    _write_sheet_headers(ws, layout)

    stock_name_idx = 1
    for i, c in enumerate(layout.basic_cols):
        if c == "stock_name":
            stock_name_idx = i + 2
            break
    ws.freeze_panes = f"{get_column_letter(stock_name_idx)}3"

    _set_column_widths(ws, display_df, layout)

    data_font = Font(name="微软雅黑", size=10, color="1D1D1F")
    stripe_fill = PatternFill(start_color="F7F8FA", end_color="F7F8FA", fill_type="solid")

    for row_idx, row in enumerate(
        dataframe_to_rows(display_df, index=False, header=False), 3,
    ):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    n_cols = len(display_df.columns)
    if ws.max_row >= 3 and n_cols > 0:
        last_col = get_column_letter(n_cols)
        data_range = f"A3:{last_col}{ws.max_row}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=["MOD(ROW(),2)=1"], fill=stripe_fill),
        )
        for row in ws.iter_rows(
            min_row=3, max_row=ws.max_row, min_col=1, max_col=n_cols,
        ):
            for cell in row:
                cell.font = data_font


def _write_sheet_merged(wb, sheet_name, df, daily_cols, weekly_cols):
    """Write merged daily+weekly DataFrame with two-row header and zebra-striped data rows."""
    display, layout = _build_merged_display_df(df, daily_cols, weekly_cols)
    _write_sheet_from_display(wb, sheet_name, display, layout)


